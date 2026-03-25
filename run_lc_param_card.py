import subprocess
import os
from datetime import datetime
import time
import signal
import re
import openpyxl

import resource

import argparse



MAX_MEM_GB = 20
MAX_MEM_BYTES = MAX_MEM_GB * 1024**3  # 1 GB = 1024^3 bytes

# python run_lc_param_card.py --patterns  --targets --output "/home/etud/Bureau/projet/indicators_lines_cols/run_class_b_10012026130521"


def limit_memory():
	soft, hard = resource.getrlimit(resource.RLIMIT_AS)
	# Only lower the soft limit; do not try to raise hard limit
	new_soft = min(MAX_MEM_BYTES, hard)
	resource.setrlimit(resource.RLIMIT_AS, (new_soft, hard))

#python run_lc.py



def import_to_excel(solver_output):
	#Step 1: Clean ANSI escape sequences
	ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
	clean_output = ansi_escape.sub('', solver_output)

	# Step 2: Extract metrics with individual fallback regexes
	effs_matches = re.findall(r'effs\s*:\s*(\d+)', clean_output)
	effs = effs_matches[-1] if effs_matches else "-"

	stop_match = re.search(r'stop\s*:\s*([A-Z_]+)', clean_output)
	stop = stop_match.group(1) if stop_match else "-"

	wck_matches = re.findall(r'wck\s*:\s*(\d+\.\d+)', clean_output)
	wck = wck_matches[-1] if wck_matches else "-"

	cpu_match = re.search(r'cpu\s*:\s*(\d+\.\d+)', clean_output)
	cpu = cpu_match.group(1) if cpu_match else "-"

	mem_match = re.search(r'mem\s*:\s*(\S+)', clean_output)
	mem = mem_match.group(1) if mem_match else "-"

	unsat_match = re.search(r'\bs\s+UNSATISFIABLE\b', clean_output)
	unsat = "Yes" if unsat_match else "No"

	wrong_dec_match = re.search(r'd\s+WRONG\s+DECISIONS\s+(\d+)', clean_output)
	wrong_dec = wrong_dec_match.group(1) if wrong_dec_match else "-"

	found_sols_match = re.search(r'd\s+FOUND\s+SOLUTIONS\s+(\d+)', clean_output)
	found_sols = found_sols_match.group(1) if found_sols_match else "-"

	complete_match = re.search(r'd\s+COMPLETE\s+EXPLORATION', clean_output)
	complete = "Yes" if complete_match else "No"

	# Step 3: Create Excel workbook
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Indicators"

	# Step 3: Create Excel workbook
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Indicators"

	headers = [
		"Pattern", "Target", "ILF",  "Run",
		"Dpts", "Effs", "Fails", "Wrgs", "Wck", "Ngds", "Sols",
		"Stop", "CPU", "Mem", "UNSAT", "WrongDec", "FoundSols", "Complete",
		"ilftime", "ilfmemo","ilfiterations"
	]
	ws.append(headers)

	# Step 4: Prepare and append row
	run_count = 1
	row = [
		pattern, target, ilf,
		f"run{run_count}",  # Run ID
		"-",                # Dpts (not extracted by fallback)
		effs,
		"-",                # Fails (not extracted by fallback)
		wrong_dec,          # Wrgs (using wrong decisions)
		wck,
		"-",                # Ngds (not extracted by fallback)
		found_sols,
		stop,
		cpu,
		mem,
		unsat,
		wrong_dec,
		found_sols,
		complete,
		time_used,
		mem_used,
		iterations
	]
	ws.append(row)

	# Step 5: Save Excel file with memory error handling
	timestamp = datetime.now().strftime("%d%m%Y%H%M%S")
	file_name = os.path.join(output_folder, f"indicators_{pattern}_{target_save}_{timestamp}.xlsx")

	try:
		wb.save(file_name)
		print(f"Data has been written to {file_name}")

	except MemoryError as mem_err:
		print(f"MemoryError when saving Excel: {mem_err}")
		# Free as much memory as possible
		import gc
		gc.collect()

		# Try saving a minimal Excel file with just an error message
		try:
			wb_minimal = openpyxl.Workbook()
			ws_min = wb_minimal.active
			ws_min.title = "Indicators"
			ws_min.append(["Status"])
			ws_min.append(["stop: ACE_MEM_LIMIT"])
			min_file_name = os.path.join(output_folder, f"indicators_{pattern}_{target_save}_{timestamp}.xlsx")
			wb_minimal.save(min_file_name)
			print(f"Minimal memory error report saved to {min_file_name}")
		except Exception as e2:
			print(f"Failed to save minimal Excel report: {e2}")

	except Exception as e:
		print(f"Error saving excel file: {e}")
		sys.exit(1)


def handle_error(e):
	stderr = e.stderr or ""
	stdout = e.stdout or ""

	# Case 1: Python MemoryError
	if "MemoryError" in stderr:
		print("OOM detected (MemoryError)")
		import_to_excel("stop: OOM ")

	# Case 2: Process killed by signal
	elif e.returncode < 0:
		try:
			sig = signal.Signals(-e.returncode)  # safe only for valid signal numbers
		except ValueError:
			sig = None

		if sig == signal.SIGINT:
			import_to_excel("stop: ACE_TIMEOUT")
		elif sig == signal.SIGKILL:
			import_to_excel("stop: OOM ")
		elif sig is not None:
			import_to_excel(f"stop: killed by {sig.name}")
		else:
			import_to_excel(f"stop: killed by unknown signal (-{e.returncode})")

	# Case 3: Non-zero exit code
	else:
		print(f"Process exited with code {e.returncode}")
		print("stdout:", stdout)
		print("stderr:", stderr)
		import_to_excel(f"stop: exit code {e.returncode}")





# for pattern in patterns:
# 	for target in targets:
# 		for common_args in common_args_sets:
# 			command = [
# 				"python", "lc5.py",
# 				"--pattern", pattern,
# 				"--target", target,
# 				"--output", output_folder,  
# 			] + common_args

# 			print(f"Running command: {' '.join(command)}")
# 			try:
# 				subprocess.run(command, check=True, timeout= 3600)
# 				time.sleep(2)
# 			except subprocess.CalledProcessError as e:
# 				print(f"Command failed: {e}")

def run_lc(patterns,targets, folder):


	common_args_sets = [
		
		[ "--timeout"],
		["--typage", "--ordre", "--timeout", "--ilf"],
		[ "--typage", "--ordre", "--timeout"],

	]
	common_args_card = [[ "--typage", "--ordre", "--card", "--timeout"]]
						#[ "--ilf","--typage", "--ordre", "--card", "--timeout"],]
						

	timestamp = datetime.now().strftime("%d%m%Y%H%M%S")
	output_folder = f"/home/etud/Bureau/projet/indicators_lines_cols/run_class_b_10012026130521"
	os.makedirs(output_folder, exist_ok=True)

	for pattern in patterns:
		for target in targets:
			for common_args in common_args_card:
				ilf = "--ilf" in common_args
				target_save= target.replace('/','')
				command = [
					"python", "lc6.py",
					"--pattern", pattern,
					"--target", target,
					"--output", output_folder,  
				] + common_args

				print(f"Running command: {' '.join(command)}")
				try:
					result = subprocess.run(
						command,
						preexec_fn=limit_memory,   
						check=True, timeout=7200
					)


				except subprocess.CalledProcessError as e:
					print(f"Process return code: {e.returncode}")
					handle_error(e)


if __name__ == "__main__":

	time_used=0
	mem_used=0
	iterations=0
	ilf= False
	target_save=''

	parser = argparse.ArgumentParser(description="Batch-run generateur.py")
	timestamp = datetime.now().strftime("%d%m%Y%H%M%S")
	default_folder = f"/home/etud/Bureau/projet/indicators_lines_cols/run_{timestamp}"

	parser.add_argument("--patterns", type=str, help="patterns as comma-separated strings")
	parser.add_argument("--targets", type=str, help="targets as comma-separated strings")
	parser.add_argument("--output", type=str, help="output folder", default=default_folder)

	args = parser.parse_args()

	patterns = args.patterns.split(",") if args.patterns else []
	targets = args.targets.split(",") if args.targets else []
	folder = args.output
	os.makedirs(folder, exist_ok=True)


	run_lc(patterns,targets, folder)