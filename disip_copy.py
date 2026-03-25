from enum import Enum, auto
from pycsp3 import *
import subprocess
import openpyxl
import resource
import shutil
import re
import argparse
from datetime import datetime 
import os
import sys
import argparse
import pygraphviz as pgv
from pathlib import Path
import inflect
import networkx as nx
import pydot
from collections import Counter
from class_label_order import LabelOrder
from class_node import Node
from ilf3 import run_with_timeout,get_label,hopcroft, build_partial_order_3, build_partial_order_4,hopcroft_multiset,dir_ILF, ILF



import os
import time
import psutil
import signal
import gc



JAVA_HEAP_GB = 8
MEMORY_LIMIT_GB = 20
HARD_LIMIT_GB = 28
monitor_threshold_gb = HARD_LIMIT_GB * 0.8




def import_to_excel(solver_output):
	#Step 1: Clean ANSI escape sequences
	ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
	clean_output = ansi_escape.sub('', solver_output)

	# Step 2: Extract metrics with individual fallback regexes
	effs_matches = re.findall(r'effs\s*:\s*(\d+)', clean_output)
	effs = effs_matches[-1] if effs_matches else "-"

	stop_match = re.search(r'stop\s*:\s*([A-Z_]+)', clean_output)
	stop = stop_match.group(1) if stop_match else "-"

	wck_matches = re.findall(r'wck\s*:\s*(\d+\.\d+)', clean_output)
	wck = wck_matches[-1] if wck_matches else "-"

	cpu_match = re.search(r'cpu\s*:\s*(\d+\.\d+)', clean_output)
	cpu = cpu_match.group(1) if cpu_match else "-"

	mem_match = re.search(r'mem\s*:\s*(\S+)', clean_output)
	mem = mem_match.group(1) if mem_match else "-"

	unsat_match = re.search(r'\bs\s+UNSATISFIABLE\b', clean_output)
	unsat = "Yes" if unsat_match else "No"

	wrong_dec_match = re.search(r'd\s+WRONG\s+DECISIONS\s+(\d+)', clean_output)
	wrong_dec = wrong_dec_match.group(1) if wrong_dec_match else "-"

	found_sols_match = re.search(r'd\s+FOUND\s+SOLUTIONS\s+(\d+)', clean_output)
	found_sols = found_sols_match.group(1) if found_sols_match else "-"

	complete_match = re.search(r'd\s+COMPLETE\s+EXPLORATION', clean_output)
	complete = "Yes" if complete_match else "No"

	# Step 3: Create Excel workbook
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Indicators"

	# Step 3: Create Excel workbook
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Indicators"

	headers = [
		"Pattern", "Target", "ILF",  "Run",
		"Dpts", "Effs", "Fails", "Wrgs", "Wck", "Ngds", "Sols",
		"Stop", "CPU", "Mem", "UNSAT", "WrongDec", "FoundSols", "Complete",
		"ilftime", "ilfmemo","ilfiterations"
	]
	ws.append(headers)

	# Step 4: Prepare and append row
	run_count = 1
	row = [
		pattern, target, ilf,
		f"run{run_count}",  # Run ID
		"-",                # Dpts (not extracted by fallback)
		effs,
		"-",                # Fails (not extracted by fallback)
		wrong_dec,          # Wrgs (using wrong decisions)
		wck,
		"-",                # Ngds (not extracted by fallback)
		found_sols,
		stop,
		cpu,
		mem,
		unsat,
		wrong_dec,
		found_sols,
		complete,
		time_used,
		mem_used,
		iterations
	]
	ws.append(row)

	# Step 5: Save Excel file with memory error handling
	timestamp = datetime.now().strftime("%d%m%Y%H%M%S")
	file_name = os.path.join(output_folder, f"indicators_{pattern}_{target}_{timestamp}.xlsx")

	try:
		wb.save(file_name)
		print(f"Data has been written to {file_name}")

	except MemoryError as mem_err:
		print(f"MemoryError when saving Excel: {mem_err}")
		# Free as much memory as possible
		import gc
		gc.collect()

		# Try saving a minimal Excel file with just an error message
		try:
			wb_minimal = openpyxl.Workbook()
			ws_min = wb_minimal.active
			ws_min.title = "Indicators"
			ws_min.append(["Status"])
			ws_min.append(["stop: ACE_MEM_LIMIT"])
			min_file_name = os.path.join(output_folder, f"indicators_{pattern}_{target}_{timestamp}.xlsx")
			wb_minimal.save(min_file_name)
			print(f"Minimal memory error report saved to {min_file_name}")
		except Exception as e2:
			print(f"Failed to save minimal Excel report: {e2}")

	except Exception as e:
		print(f"Error saving excel file: {e}")
		sys.exit(1)








def get_process_tree_pids(root_pid):
	"""Return root PID + all descendant PIDs."""
	try:
		root_proc = psutil.Process(root_pid)
		return [root_pid] + [p.pid for p in root_proc.children(recursive=True)]
	except psutil.NoSuchProcess:
		return []

def _kill_process_group(root_pid: int):
	"""Kill the process group for root_pid on Linux."""
	try:
		pgid = os.getpgid(root_pid)
		os.killpg(pgid, signal.SIGKILL)
		print(f"[Monitor] Killed process group {pgid}")
	except ProcessLookupError:
		print(f"[Monitor] Process group for PID {root_pid} no longer exists.")
	except PermissionError as e:
		print(f"[Monitor] Permission denied killing PGID for {root_pid}: {e}")
	except Exception as e:
		print(f"[Monitor] Error killing PGID for {root_pid}: {e}")
		# Fallback: kill root PID only
		try:
			os.kill(root_pid, signal.SIGKILL)
			print(f"[Monitor] Killed root PID {root_pid}")
		except ProcessLookupError:
			pass
		except Exception as e2:
			print(f"[Monitor] Could not kill root PID {root_pid}: {e2}")

def monitor_and_kill_multiple(root_pid, mem_threshold_gb, timeout_sec=None, poll=0.2):
	"""
	Monitor total RSS memory usage of root process + children,
	kill process group if time/memory threshold exceeded.
	"""
	start_time = time.time()
	TIMEOUT_EVENT.clear()
	MEMORY_EVENT.clear()

	while True:
		elapsed = time.time() - start_time

		# Refresh PIDs every loop
		all_pids = get_process_tree_pids(root_pid)

		# Timeout check
		if timeout_sec is not None and elapsed >= timeout_sec and not TIMEOUT_EVENT.is_set():
			print(f"[Monitor] Time limit {timeout_sec}s reached. Killing process group of {root_pid}")
			TIMEOUT_EVENT.set()
			_kill_process_group(root_pid)
			break

		# Memory check
		total_mem_gb = 0.0
		for pid in all_pids:
			try:
				proc = psutil.Process(pid)
				total_mem_gb += proc.memory_info().rss / (1024**3)
			except (psutil.NoSuchProcess, psutil.AccessDenied):
				pass

		if total_mem_gb >= mem_threshold_gb and not MEMORY_EVENT.is_set():
			print(f"[Monitor] Memory limit exceeded ({total_mem_gb:.2f} GB >= {mem_threshold_gb:.2f} GB). Killing process group of {root_pid}")
			MEMORY_EVENT.set()
			_kill_process_group(root_pid)
			break

		# Stop if root process gone
		if not psutil.pid_exists(root_pid):
			break

		time.sleep(poll)


import subprocess
import threading
import psutil
import os
import signal
import time
import select
def run_ace_with_monitor(command, mem_limit_gb, timeout, output_path):
	"""
	Launch ACE with a watchdog that enforces memory & timeout limits.
	Detects Python MemoryError during reading and logs to Excel.
	"""
	process = subprocess.Popen(
		command,
		stdout=subprocess.PIPE,
		stderr=subprocess.PIPE,
		text=True,
		bufsize=1,
		universal_newlines=True,
		start_new_session=True
	)

	solver_pid = process.pid

	# Flags for monitor
	stop_event = threading.Event()
	memory_exceeded = threading.Event()
	timeout_exceeded = threading.Event()
	python_oom = threading.Event()  # declare before loop

	# Watchdog thread
	def watchdog():
		proc = psutil.Process(solver_pid)
		start_time = time.time()
		while not stop_event.is_set():
			if time.time() - start_time > timeout:
				timeout_exceeded.set()
				try:
					import_to_excel('stop: ACE_timeout')
					os.killpg(os.getpgid(solver_pid), signal.SIGKILL)
					print(f"[Watchdog] Timeout -> killed {solver_pid}")
				except ProcessLookupError:
					pass
				break
			try:
				mem_used_gb = proc.memory_info().rss / (1024 ** 3)
				if mem_used_gb > mem_limit_gb:
					memory_exceeded.set()
					try:
						os.killpg(os.getpgid(solver_pid), signal.SIGKILL)
						print(f"[Watchdog] Memory limit exceeded -> killed {solver_pid}")
					except ProcessLookupError:
						pass
					break
			except psutil.NoSuchProcess:
				break
			time.sleep(0.1)

	threading.Thread(target=watchdog, daemon=True).start()

	stdout_lines = []
	stderr_lines = []
	fds = [process.stdout.fileno(), process.stderr.fileno()]

	try:
		while process.poll() is None and not (memory_exceeded.is_set() or timeout_exceeded.is_set()):
			ready_fds, _, _ = select.select(fds, [], [], 0.1)
			for fd in ready_fds:
				if fd == process.stdout.fileno():
					line = process.stdout.readline()
					if line:
						stdout_lines.append(line)
				elif fd == process.stderr.fileno():
					line = process.stderr.readline()
					if line:
						stderr_lines.append(line)

		if process.poll() is None:
			os.killpg(os.getpgid(solver_pid), signal.SIGKILL)

		stdout_lines += process.stdout.read().splitlines()
		stderr_lines += process.stderr.read().splitlines()


	finally:
		process.stdout.close()
		process.stderr.close()
		stop_event.set()

	stdout_text = "\n".join(stdout_lines)
	stderr_text = "\n".join(stderr_lines)

	
	return stdout_text, stderr_text, solver_pid



def load(instance, graph_file, id, node_dict={}):
	
	# Load graph
	graph = nx.drawing.nx_agraph.read_dot(graph_file).to_directed()

	# Plot input graph
	pydot.graph_from_dot_file(graph_file)[0].write_png(f'res/{Path(os.path.basename(graph_file)).stem}.png')

	#print("Parsed nodes:", graph.nodes(data=True))

	#edges = graph.edges
	edges = [(u, v, d['angle']) for u, v, d in graph.edges(data=True) if 'angle' in d]
	labels = nx.get_node_attributes(graph, 'label')
	angles = nx.get_edge_attributes(graph, 'angle')
	graphs = pydot.graph_from_dot_file(graph_file)
	pydot_graph = graphs[0]
	nodes = pydot_graph.get_nodes()
	#print("Total nodes in PyDot:", len(nodes))
	#print("Node labels:", labels)

	# Set instance data
	instance[f'NV_{id}'] = graph.number_of_nodes()
	instance[f'NE_{id}'] = graph.number_of_edges()
	instance[f'V{id}'] = [label for label in labels.values()]
	instance[f'E{id}'] = list()
	instance[f'A{id}'] = list()

	for n1, n2, a in edges:
		instance[f'E{id}'].append((int(n1), int(n2)))
		key = (n1, n2, 0) if (n1, n2, 0) in angles else (n1, n2)
		instance[f'A{id}'].append('Horizontal' if angles[key] == '0' else 'Vertical')



def process_mapping( instance, mapping_n, mapping, target,pattern):

	print(f'{mapping_n}.\t[{", ".join(instance["V1"])}] -> [{", ".join(list(map(lambda i: instance["V2"][i], mapping)))}]')

	# Plot mapping
	graph = pydot.graph_from_dot_file(f'dat/{target}.dot')[0]
	for v1 in range(instance['NV_1']):
		graph.get_node(str(mapping[v1]))[0].set_label(instance['V1'][v1])
		graph.get_node(str(mapping[v1]))[0].set_style('filled')
	graph.write_png(f'res/{pattern}_in_{target}_{mapping_n}.png')


# Angle of the edge of a graph
class Angle(str, Enum):
	Horizontal = 'Horizontal'
	Vertical = 'Vertical'

def domain_n(i, NV_2):
	if ilf:
		return node_dict[f'{i}p']
	else:
		return range(NV_2 )
def model(

	# =============== DATA
	
	#dom: list[str],
	# Pattern graph G1 = (V1, E1), with |V1| = NV_1, |E1| = NV_1, and A1 the angle of each edges in E1
	NV_1: int, 
	NE_1: int, 
	V1: list[str],
	E1: list[list[int]],
	A1: list[Angle],

	# Target graph G2 = (V2, E2), with |V2| = NV_2, |E2| = NV_2, and A2 the angle of each edges in E2
	NV_2: int, 
	NE_2: int, 
	V2: list[str],
	E2: list[list[int]],
	A2: list[Angle]
	):


	# =============== VARIABLE

	# Mapping nodes from graph 1 to graph 2
	#if ilf:
	I = VarArray(size=NV_1, dom=lambda i: domain_n(i, NV_2))
	# else:
	# 	I = VarArray(size=NV_1, dom=range(NV_2 ))
	#T = {( i , j ) for i , j in E2 } | {( j , i ) for i , j in E2 }

	# =============== CONSTRAINT

	satisfy(

		AllDifferent(I),

		[
		Exist(
		(I[E1[e1][0]] == E2[e2][0]) & (I[E1[e1][1]] == E2[e2][1])
		for e2 in range(NE_2) if A1[e1]==A2[e2]
		) for e1 in range(NE_1)
		],

		[
		#((disjunction(I[n1] == node_dict[f'{n1+1}p'][i % len(node_dict[f'{n1+1}p'])])for i in range(len(node_dict[f'{n1+1}p'])) )for n1 in range(NV_1))
		#((I[n1] in node_dict[f'{n1+1}p'])for n1 in range(NV_1))
		]
	)
	# for i in range(NV_1):
	# 	print(I[i].dom)
	# print(posted())
	return I


def set_memory_limit_gb(gb_soft_limit, gb_hard_limit=16):
	soft = int(gb_soft_limit * 1024**3)
	hard = int(gb_hard_limit * 1024**3)
	resource.setrlimit(resource.RLIMIT_AS, (soft, hard))




if __name__ == '__main__':


	set_memory_limit_gb(MEMORY_LIMIT_GB, HARD_LIMIT_GB)

	python_pid = os.getpid()  # current Python process




	time_used=0
	mem_used=0
	iterations=0
	ilf_empty_domain = False
	ilf_reached_timeout= False
	timestamp = datetime.now().strftime("%d%m%Y%H%M%S")

	#python disip.py --pattern ladder  --target gt-06122025092950 --ilf --ordre --typage

	# parser = argparse.ArgumentParser(description='Process graph matching parameters.')
	# parser.add_argument('--pattern', type=str, required=True, help='Name of the pattern file (without .dot extension)')
	# parser.add_argument('--target', type=str, required=True, help='Name of the target file (without .dot extension)')
	# parser.add_argument('--ilf', action='store_true', help='Enable ILF option')
	# parser.add_argument('--ordre', action='store_true', help='Enable order option')
	# parser.add_argument('--typage', action='store_true', help='Enable typing option')
	# parser.add_argument("--output", type=str, required=True) 

	# args = parser.parse_args()

	# pattern = args.pattern
	# target = args.target
	# ilf = args.ilf
	# ordre = args.ordre
	# typage = args.typage
	# output_folder= args.output


	pattern= "pan_apositions-09212025143931"
	target= "apositions-09232025154832"
	ilf = False
	output_folder= "/home/etud/Bureau/projet/indicators_pos/"


	

	solver = ACE
	# Go to parent directory
	os.chdir(os.path.dirname(os.path.realpath(__file__)) + '/..')
	#print(os.path.dirname(os.path.realpath(__file__))+ '/..') 
	# Create a solving instance of the model
	instance = dict()


	Gp = nx.drawing.nx_agraph.read_dot(f'dat/{pattern}.dot')
	Gt = nx.drawing.nx_agraph.read_dot(f'dat/{target}.dot')
	Gp = nx.relabel_nodes(Gp, {node: f"{node}p" for node in Gp.nodes()})
	
	if ilf:
		node_dict,iterations,  time_used, mem_used = run_with_timeout(7200 , dir_ILF, Gp, Gt) #timeout 2 hrs
		
		if node_dict == "TIMEOUT":
			ilf_reached_timeout = True
		elif node_dict == "ERROR":
			print("ILF crashed or raised exception")
		elif node_dict is None:
			ilf_empty_domain = True
	else:
		node_dict = {}

	if ilf_reached_timeout:
		print("ilf_reached_timeout")
		import_to_excel('stop: ILF_TIMEOUT') #fills Stop header with ILF_TIMEOUT
	elif ilf_empty_domain:
		import_to_excel('stop: EMPTY_DOMAIN_ILF')
###solve#####################################################################################


	else:
		# Load data to the instance
		instance = dict ()
		load(instance, f'dat/{pattern}.dot', 1, node_dict)
		load(instance, f'dat/{target}.dot', 2)

		I = model(**instance)

		
		timeout = 3600

		compile()  # generates lc6.xml by default
		result = solve(sols=ALL)
		xml_filename = f"disipcopy_{pattern}_{target}_{timestamp}.xml"

		
		shutil.move("disip_copy.xml", xml_filename)

		# After ILF is done
		del Gp, Gt, node_dict, instance  # delete big objects
		gc.collect()  # force garbage collection


		ace_jar_path = "/home/etud/Bureau/projet/lib/python3.11/site-packages/pycsp3/solvers/ace/ACE-2.3.jar"



		# Where to save solver output directly
		solver_output_path = os.path.join(output_folder, "solver_output.txt")

		if os.path.exists(xml_filename):
			# command = ["java",   "-Xmx5g", "-jar", ace_jar_path, xml_filename, "-s=all", f"-t={timeout}s", "-trace", "-ev"]
			
			try:
				# ---- start ACE
				command = [
					"java",
					f"-Xmx{JAVA_HEAP_GB}g",
					"-jar", ace_jar_path,
					xml_filename,
					"-s=all",
					f"-t={timeout}s",  # optional; keep or remove; our monitor enforces hard timeout anyway
					"-trace",
					"-ev"
				]

				# Run ACE with non-blocking monitor
				stdout, stderr, solver_pid = run_ace_with_monitor(
				command=command,
				mem_limit_gb=MEMORY_LIMIT_GB,
				timeout=timeout,
				output_path=solver_output_path
			)
				solver_output = stdout or ""
				num_solutions = n_solutions() or 0  # Default to 0 if None

				if solver_output:
						import_to_excel(stdout or "")
						print("ACE completed successfully.")
				if num_solutions > 0:
					message=f'Patterns {pattern} in  {target}:'
					message += ' with ilf' if ilf else ' without ilf'
					print(message)
					#for solution_i in range(n_solutions()):
					for solution_i in range(min(3, n_solutions())):
						process_mapping( instance, solution_i + 1, values(I, sol=solution_i),target,pattern)
		

			except subprocess.TimeoutExpired:
				print(f"ACE process timed out after {timeout} seconds.")
				import_to_excel('stop: ACE_TIMEOUT')
				for pid in [python_pid, solver_pid]:
					try:
						os.kill(pid, signal.SIGKILL)
					except ProcessLookupError:
						pass

			except Exception as e:
				print(f"Error occurred: {e}")
				for pid in [python_pid]:
					try:
						os.kill(pid, signal.SIGKILL)
					except ProcessLookupError:
						pass




		else:
			print("Error: disip.xml file not generated. Check your model definition.")
		


gc.collect()

	