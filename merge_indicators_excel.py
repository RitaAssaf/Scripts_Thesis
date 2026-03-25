import os
import openpyxl
import argparse

# Example usage:
# python merge_indicators_excel.py --file indicators_pos --keyword gt; python organize_excel.py --file merged_indicators

# Argument parsing
parser = argparse.ArgumentParser(description='Merge last rows from Excel files in a specified directory.')
parser.add_argument('--file', type=str, required=True, help='Name of the directory inside /projet/ containing Excel files')
parser.add_argument('--keyword', type=str, default="", help='Optional keyword to filter filenames (e.g., "gt")')
args = parser.parse_args()

file = args.file
keyword = args.keyword

# Paths
input_dir = f"/home/etud/Bureau/projet/{file}/"
output_file = os.path.join(input_dir, f"merged_indicators{f'_{keyword}' if keyword else ''}.xlsx")

# Merge data setup
merged_rows = []
header_written = False
header = None

# Loop through Excel files
for filename in os.listdir(input_dir):
    if filename.endswith(".xlsx") and (keyword in filename):
        filepath = os.path.join(input_dir, filename)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active

            # Get non-empty rows
            rows = [row for row in ws.iter_rows(values_only=True) if any(cell is not None for cell in row)]
            if not rows:
                continue

            # Capture header once
            if not header_written:
                header = rows[0]
                header_written = True

            # Get last row
            last_row = rows[-1]
            merged_rows.append(last_row)

        except Exception as e:
            print(f"Error reading {filename}: {e}")

# Write output file
if merged_rows:
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Merged Indicators"

    if header:
        ws_out.append(header)

    for row in merged_rows:
        ws_out.append(row)

    try:
        wb_out.save(output_file)
        print(f"Merged file saved as: {output_file}")
    except Exception as e:
        print(f"Error saving merged file: {e}")
else:
    print("No valid data found to merge.")
