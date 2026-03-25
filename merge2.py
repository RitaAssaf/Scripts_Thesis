import os
import openpyxl
import pandas as pd
import argparse
from datetime import datetime 


# Usage:
# python merge2.py --folder indicators_pos --subfolder run_10122025143247

# Argument parsing
parser = argparse.ArgumentParser(description='Merge last rows from Excel files and organize by pattern.')
parser.add_argument('--folder', type=str, required=True, help='Name of the folder inside /projet/ containing Excel files')
parser.add_argument('--subfolder', type=str, required=True, help='Name of the folder inside /projet/ containing Excel files')

parser.add_argument('--keyword', type=str, default="", help='Optional keyword to filter filenames')
args = parser.parse_args()

folder = args.folder
subfolder = args.subfolder

keyword = args.keyword

# Paths
base_path = f"/home/etud/Bureau/projet/{folder}/{subfolder}/"
timestamp = datetime.now().strftime("%d%m%Y%H%M%S")

merged_filename = f"merged_indicators{f'_{keyword}' if keyword else ''}.xlsx"
merged_file_path = os.path.join(base_path, merged_filename)
organized_file_path = os.path.join(base_path, f"results.xlsx")

# Step 1: Merge
merged_rows = []
header_written = False
header = None

for filename in os.listdir(base_path):
    if filename.endswith(".xlsx") and (keyword in filename):
        filepath = os.path.join(base_path, filename)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            rows = [row for row in ws.iter_rows(values_only=True) if any(cell is not None for cell in row)]
            if not rows:
                continue

            if not header_written:
                header = rows[0]
                header_written = True

            last_row = rows[-1]
            merged_rows.append(last_row)
        except Exception as e:
            print(f"Error reading {filename}: {e}")

if merged_rows:
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Merged Indicators"

    if header:
        ws_out.append(header)
    for row in merged_rows:
        ws_out.append(row)

    try:
        wb_out.save(merged_file_path)
        print(f"Merged file saved as: {merged_file_path}")
    except Exception as e:
        print(f"Error saving merged file: {e}")
else:
    print("No valid data found to merge.")
    exit(0)

# Step 2: Organize + Remove duplicates
try:
    df = pd.read_excel(merged_file_path)

    # Remove duplicate rows
    df = df.drop_duplicates()

    # Sort by Target, then Pattern
    df_sorted = df.sort_values(by=['Target', 'Pattern'], na_position="last")

    df_sorted.to_excel(organized_file_path, index=False)
    print(f"Organized file saved as: {organized_file_path}")
except Exception as e:
    print(f"Error during organizing step: {e}")
