import os
import openpyxl

import argparse
#python merge_indicators.py --file indicators_pos; python organize_excel.py --file merged_indicators
parser = argparse.ArgumentParser(description='Process graph matching parameters.')
parser.add_argument('--file', type=str, required=True, help='Name of the  file (without .dot extension)')
args = parser.parse_args()	
file = args.file

#python merge_indicators.py; python organize_excel.py --file merged_indicators
# Directory containing Excel files
input_dir = f"/home/etud/Bureau/projet/{file}/"
output_file = os.path.join(input_dir, "merged_indicators.xlsx")

# Collect last rows
merged_rows = []
header_written = False
header = None

# Loop through each Excel file in the directory
for filename in os.listdir(input_dir):
    if filename.endswith(".xlsx"):
        filepath = os.path.join(input_dir, filename)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active

            # Get all rows and remove completely empty ones
            rows = [row for row in ws.iter_rows(values_only=True) if any(cell is not None for cell in row)]

            if not rows:
                continue

            # Save header from the first file
            if not header_written:
                header = rows[0]
                header_written = True

            # Get last non-empty row (excluding header)
            last_row = rows[-1]
            merged_rows.append(last_row)

        except Exception as e:
            print(f"Error reading {filename}: {e}")

# Write merged data to a new Excel file
if merged_rows:
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Merged Indicators"

    if header:
        ws_out.append(header)

    for row in merged_rows:
        ws_out.append(row)

    try:
        wb_out.save(output_file)
        print(f"Merged file saved as: {output_file}")
    except Exception as e:
        print(f"Error saving merged file: {e}")
else:
    print("No valid data found to merge.")
