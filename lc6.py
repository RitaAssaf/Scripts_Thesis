from enum import Enum, auto
from pycsp3 import *
import subprocess
import os
import sys
import argparse
import gc
import pygraphviz as pgv
from pathlib import Path
import inflect
import networkx as nx
import pydot
import re
import openpyxl
import subprocess
import psutil
import threading
import csv
import time
import shutil
from collections import Counter
from class_label_order import LabelOrder

from datetime import datetime 
from ilf3 import run_with_timeout,get_label,hopcroft, build_partial_order_3, build_partial_order_4,hopcroft_multiset,dir_ILF, ILF
from itertools import combinations


import resource

#config#################################################################
# JAVA_HEAP_GB = 2
# MEMORY_LIMIT_GB = 6
# HARD_LIMIT_GB = 16
# monitor_threshold_gb = MEMORY_LIMIT_GB * 0.8 

JAVA_HEAP_GB = 8
MEMORY_LIMIT_GB = 20
HARD_LIMIT_GB = 28
monitor_threshold_gb = HARD_LIMIT_GB * 0.8



def set_memory_limit_gb(gb_soft_limit, gb_hard_limit=16):
	soft = int(gb_soft_limit * 1024**3)
	hard = int(gb_hard_limit * 1024**3)
	resource.setrlimit(resource.RLIMIT_AS, (soft, hard))



def read_graph_from_file(filename):
	G = nx.Graph()

	with open(filename, "r") as f:
		lines = f.readlines()

	# Number of nodes (not strictly needed, but kept for clarity)
	n = int(lines[0].strip())

	# First pass: create nodes with attributes
	for idx, line in enumerate(lines[1:]):
		parts = line.strip().split(";")

		neighbors_str = parts[0].strip()
		card = parts[1].strip()
		node_type = parts[2].strip()
		preceds_str = parts[3].strip()
		realname = parts[4].strip()

		preceds = preceds_str.split() if preceds_str else []

		G.add_node(
			idx,
			card=card,
			type=node_type,
			realname=realname,
			preceds=preceds
		)

	# Second pass: add edges
	for idx, line in enumerate(lines[1:]):
		parts = line.strip().split(";")
		neighbors_str = parts[0].strip()

		if neighbors_str:
			neighbors = neighbors_str.split()
			for n_idx in neighbors:
				G.add_edge(idx, int(n_idx))

	return G





def load(instance, graph, id, node_dict={}):


	edges = graph.edges  # if multiedgeview add strict to graphiz
	lenedges = len(edges)
	labels = nx.get_node_attributes(graph, 'label')
	types = nx.get_node_attributes(graph, 'type')
	preceds = nx.get_node_attributes(graph, 'preceds')  # Extract preceds attribute
	cards = nx.get_node_attributes(graph, 'card')
	realnames= nx.get_node_attributes(graph, 'realname')
	# graphs = pydot.graph_from_dot_file(graph_file)
	# pydot_graph = graphs[0]

	# Create a mapping of node labels to indices
	label_to_index = {label: int(node) for node, label in labels.items()}

	# Build preced_t list of lists based on preceds attribute
	preced_t = []
	for i in range(len(labels)):
		node_name = str(i)
		if node_name in preceds:
			# Convert preceds node labels to their corresponding indices
			preced_indices = [label_to_index[succ] for succ in preceds[node_name].split(',') if succ]
			preced_t.append(preced_indices)
		else:
			preced_t.append([])  # No preceds for this node


	instance[f'preced_t_{id}'] = preced_t  

	# Set instance data
	instance[f'NV_{id}'] = graph.number_of_nodes()
	instance[f'NE_{id}'] = graph.number_of_edges()
	instance[f'V{id}'] = [name for name in labels.values()]
	instance[f'type{id}'] = [type for type in types.values()]
	instance[f'E{id}'] = list()

	instance[f'card{id}'] = [int(c) for c in cards.values()]
	instance[f'max_card_{id}'] = max(instance[f'card{id}'])
	for n1, n2 in edges:
		instance[f'E{id}'].append((int(n1), int(n2)))
	instance[f'A{id}'] = list()
	instance[f'realnames{id}'] = [name for name in realnames.values()]




def import_to_excel(solver_output):
	#Step 1: Clean ANSI escape sequences
	ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
	clean_output = ansi_escape.sub('', solver_output)

	# Step 2: Extract metrics with individual fallback regexes
	effs_matches = re.findall(r'effs\s*:\s*(\d+)', clean_output)
	effs = effs_matches[-1] if effs_matches else "-"

	stop_match = re.search(r'stop\s*:\s*([A-Z_]+)', clean_output)
	stop = stop_match.group(1) if stop_match else "-"

	wck_matches = re.findall(r'wck\s*:\s*(\d+\.\d+)', clean_output)
	wck = wck_matches[-1] if wck_matches else "-"

	cpu_match = re.search(r'cpu\s*:\s*(\d+\.\d+)', clean_output)
	cpu = cpu_match.group(1) if cpu_match else "-"

	mem_match = re.search(r'mem\s*:\s*(\S+)', clean_output)
	mem = mem_match.group(1) if mem_match else "-"

	fails_matches = re.findall(r'fails\s*:\s*(\d+)', clean_output)
	fails = fails_matches[-1] if fails_matches else "-"


	unsat_match = re.search(r'\bs\s+UNSATISFIABLE\b', clean_output)
	unsat = "Yes" if unsat_match else "No"

	wrong_dec_match = re.search(r'd\s+WRONG\s+DECISIONS\s+(\d+)', clean_output)
	wrong_dec = wrong_dec_match.group(1) if wrong_dec_match else "-"

	found_sols_match = re.search(r'd\s+FOUND\s+SOLUTIONS\s+(\d+)', clean_output)
	found_sols = found_sols_match.group(1) if found_sols_match else "-"

	complete_match = re.search(r'd\s+COMPLETE\s+EXPLORATION', clean_output)
	complete = "Yes" if complete_match else "No"

	# Step 3: Create Excel workbook
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Indicators"

	# Step 3: Create Excel workbook
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Indicators"

	headers = [
		"Pattern", "Target", "ILF", "Order", "Typing", "Card", "Run",
		"Dpts", "Effs", "Fails", "Wrgs", "Wck", "Ngds", "Sols",
		"Stop", "CPU", "Mem", "UNSAT", "WrongDec", "FoundSols", "Complete",
		"ilftime", "ilfmemo","ilfiterations"
	]
	ws.append(headers)

	# Step 4: Prepare and append row
	run_count = 1
	row = [
		pattern, target, ilf, ordre, typage, card,
		f"run{run_count}",  # Run ID
		"-",                # Dpts (not extracted by fallback)
		effs,
		fails,
		wrong_dec,          # Wrgs (using wrong decisions)
		wck,
		"-",                # Ngds (not extracted by fallback)
		found_sols,
		stop,
		cpu,
		mem,
		unsat,
		wrong_dec,
		found_sols,
		complete,
		time_used,
		mem_used,
		iterations
	]
	ws.append(row)

	# Step 5: Save Excel file with memory error handling
	timestamp = datetime.now().strftime("%d%m%Y%H%M%S")
	file_name = os.path.join(output_folder, f"indicators_{pattern}_{target_save}_{timestamp}.xlsx")

	try:
		wb.save(file_name)
		print(f"Data has been written to {file_name}")

	except MemoryError as mem_err:
		print(f"MemoryError when saving Excel: {mem_err}")
		# Free as much memory as possible
		import gc
		gc.collect()

		# Try saving a minimal Excel file with just an error message
		try:
			wb_minimal = openpyxl.Workbook()
			ws_min = wb_minimal.active
			ws_min.title = "Indicators"
			ws_min.append(["Status"])
			ws_min.append(["stop: ACE_MEM_LIMIT"])
			min_file_name = os.path.join(output_folder, f"indicators_{pattern}_{target_save}_{timestamp}.xlsx")
			wb_minimal.save(min_file_name)
			print(f"Minimal memory error report saved to {min_file_name}")
		except Exception as e2:
			print(f"Failed to save minimal Excel report: {e2}")

	except Exception as e:
		print(f"Error saving excel file: {e}")
		sys.exit(1)



def import_to_csv(file_stream):
	for line in file_stream:
		# parse and write to CSV incrementally
		pass


def get_node_label(G, label):
	#return [n for n, d in G.nodes(data=True) if d.get("label") == "Person"][0]
	return {n.get_name(): n for n in G.get_nodes()}[label] #graphiz file


def process_mapping( instance, mapping_n, mapping, target,pattern):

	print(f'{mapping_n}.\t[{", ".join(instance["V1"])}] -> [{", ".join(list(map(lambda i: instance["realnames2"][i], mapping)))}]')

	# Plot mapping
	graph = pydot.graph_from_dot_file(f'lcres/{target}.dot')[0]
	for v1 in range(instance['NV_1']):#
		# prend le label du noeud target qui est dans l'array V2 a l'indice solution trouvé par le solveur
		#donc le solveur retourne array des indices des noeuds de la solution 
		node=instance['V2'][mapping[v1]] 
		get_node_label(graph, node).set_label(instance['V1'][v1])
		#graph.get_node(node)[0].set_label(instance['V1'][v1])
		get_node_label(graph, node).set_style('filled')
		graph.write_png(f'res/{pattern}_in_{target_save}_{mapping_n}.png')


# Angle of the edge of a graph
class Angle(str, Enum):
	Horizontal = 'Horizontal'
	Vertical = 'Vertical'

def domain_n(i, NV_2):
	
	indices_l= [i for i, type in enumerate(instance['type2'] ) if type == 'l']

	indices_c= [i for i, type in enumerate(instance['type2'] ) if type == 'c']
	#si le type du noeud patterne est l donc il aura les indices(qui sont les noms aussi) des noeuds l
	
	if not ilf and not typage:	
		return range(NV_2 )
	elif not ilf and typage:	
		if instance['type1'][i]== 'l':
			return indices_l
		else: return indices_c
	elif ilf and not typage:	
		return node_dict[f'{i}p']
	elif ilf and typage:
		ilf_domain= node_dict[f'{i}p']
		if instance['type1'][i]== 'l':
			return list(set(ilf_domain) & set(indices_l))
		else: return list(set(ilf_domain) & set(indices_c))
	

def dom_preced(i: int, j: int, target):
	if target:
		if 0 <= i < len(preced_t_graph) and 0 <= j < len(preced_t_graph[i]):
			return [preced_t_graph[i][j]] 
		return [-1] 
	else:
		if 0 <= i < len(preced_p_graph) and 0 <= j < len(preced_p_graph[i]):
			return [preced_p_graph[i][j]] 
		return [-1] 

def fill_matrix(lst_of_lsts,length):
	
	result = [sublist + [-1] * (length - len(sublist)) for sublist in lst_of_lsts]  # Fill shorter lists with -1
	
	return result


#Function model takes as parameter nodes,edges and retrun var array
#builds the model but returns only I

def model(

	# =============== DATA
	
	# Pattern graph G1 = (V1, E1), with |V1| = NV_1, |E1| = NV_1, and A1 the angle of each edges in E1
	NV_1: int, 
	NE_1: int, 
	V1: list[str],
	E1: list[list[int]],
	A1: list[Angle],
	type1:list[str],
	preced_t_1:list[list[int]],
	
	# Target graph G2 = (V2, E2), with |V2| = NV_2, |E2| = NV_2, and A2 the angle of each edges in E2
	NV_2: int, 
	NE_2: int, 
	V2: list[str],
	E2: list[list[int]],
	A2: list[Angle],
	type2:list[str],
	preced_t_2:list[list[int]],
	card1=[],
	card2=[],
	realnames1=[],
	realnames2=[],
	max_card_1=0,
	max_card_2=0
):


	# =============== VARIABLE


	I = VarArray(size=NV_1, dom=lambda i: domain_n(i, NV_2))

	if card:
		cardinality_p= VarArray(size= NV_1, dom=range(max_card_1+1))
		cardinality_t=VarArray(size= NV_2, dom=range(max_card_2+1))
		
		# cardinality_p= VarArray(size= NV_1, dom=range(15))
		# cardinality_t=VarArray(size= NV_2, dom=range(15))
	
	T = {( i , j ) for i , j in E2 } | {( j , i ) for i , j in E2 }

	# =============== CONSTRAINT
	satisfy(

		AllDifferent(I),

		[( I [ i ] , I [ j ]) in T for (i , j ) in E1 ] ,

	)
	
	if ordre : 
		satisfy (
			[( I [ i ] , I [ j ]) in pairs_t for (i , j ) in pairs_p ] 
		)

	if card :
		satisfy (
			[cardinality_p[n] == card_p[n] for n in range(NV_1)], 
			[cardinality_t[n] == card_t[n] for n in range(NV_2)],
			[cardinality_t[I[i]] == cardinality_p[i] for i in range(NV_1)]
		)
	# for i in range(NV_1):
	# 	print(I[i].dom)
	#print(posted())
	return I


MEMORY_KILLED = False
MEMORY_KILLED_REASON = None


def monitor_and_kill(pid, threshold_gb=6, interval=5):
	"""Monitor a single process (and its children) memory and kill if exceeds threshold."""
	global MEMORY_KILLED
	try:
		parent = psutil.Process(pid)
		while parent.is_running():
			# Calc6ulate total memory usage (parent + children) in GB
			mem_used_gb = parent.memory_info().rss / (1024 ** 3)
			for child in parent.children(recursive=True):
				mem_used_gb += child.memory_info().rss / (1024 ** 3)

			if mem_used_gb > threshold_gb:
				MEMORY_KILLED = True
				print(f"[WARNING] Memory exceeded {threshold_gb} GB (used {mem_used_gb:.2f} GB) — Killing processes...")
				import_to_excel('stop: ACE_MEM_OUT')
				with open("memory_log.txt", "a") as log_file:
					log_file.write(f"[{datetime.now()}] Killed process due to memory limit ({mem_used_gb:.2f} GB)\n")

				# Terminate children first
				for child in parent.children(recursive=True):
					try:
						child.terminate()
					except psutil.NoSuchProcess:
						pass

				parent.terminate()
				try:
					parent.wait(timeout=5)
				except psutil.TimeoutExpired:
					parent.kill()
				break

			time.sleep(interval)
	except psutil.NoSuchProcess:
		pass


# def monitor_and_kill_multiple(pids, threshold_gb=6, interval=5):
# 	"""Start a separate monitor thread for each PID in a list."""
# 	import threading

# 	threads = []
# 	for pid in pids:
# 		if isinstance(pid, int):
# 			t = threading.Thread(target=monitor_and_kill, args=(pid, threshold_gb, interval), daemon=True)
# 			t.start()
# 			threads.append(t)
# 		else:
# 			print(f"Warning: PID {pid} is not an integer and will be skipped")

import time, os, signal, psutil, threading

# global events instead of plain booleans
TIMEOUT_EVENT = threading.Event()
MEMORY_EVENT = threading.Event()


def _kill_process_group(root_pid: int):
	try:
		pgid = os.getpgid(root_pid)
		os.killpg(pgid, signal.SIGKILL)  # kill the whole Java group
	except Exception:
		# fallback: kill just the root pid
		try:
			os.kill(root_pid, signal.SIGKILL)
		except Exception:
			pass

def monitor_and_kill_multiple(root_pid, pids, mem_threshold_gb, timeout_sec=None, poll=0.2):
	start_time = time.time()
	# clear events at the start (in case of re-runs)
	TIMEOUT_EVENT.clear()
	MEMORY_EVENT.clear()

	while True:
		elapsed = time.time() - start_time

		# TIMEOUT
		if timeout_sec is not None and elapsed >= timeout_sec and not TIMEOUT_EVENT.is_set():
			print(f"[Monitor] Time limit {timeout_sec}s reached. Killing process group of {root_pid}")
			TIMEOUT_EVENT.set()
			_kill_process_group(root_pid)
			break

		# MEMORY
		total_mem_gb = 0.0
		for pid in pids:
			try:
				proc = psutil.Process(pid)
				total_mem_gb += proc.memory_info().rss / (1024**3)
			except psutil.NoSuchProcess:
				pass

		if total_mem_gb >= mem_threshold_gb and not MEMORY_EVENT.is_set():
			print(f"[Monitor] Memory limit exceeded ({total_mem_gb:.2f} GB >= {mem_threshold_gb:.2f} GB). Killing process group of {root_pid}")
			MEMORY_EVENT.set()
			_kill_process_group(root_pid)
			break

		# stop monitoring if the root process is gone
		if not psutil.pid_exists(root_pid):
			break

		time.sleep(poll)


def get_process_tree_pids(root_pid):
	try:
		root_proc = psutil.Process(root_pid)
		return [root_pid] + [p.pid for p in root_proc.children(recursive=True)]
	except psutil.NoSuchProcess:
		return []


import os
import time
import psutil
import signal

def get_process_tree_pids(root_pid):
	"""Return root PID + all descendant PIDs."""
	try:
		root_proc = psutil.Process(root_pid)
		return [root_pid] + [p.pid for p in root_proc.children(recursive=True)]
	except psutil.NoSuchProcess:
		return []

def _kill_process_group(root_pid: int):
	"""Kill the process group for root_pid on Linux."""
	try:
		pgid = os.getpgid(root_pid)
		os.killpg(pgid, signal.SIGKILL)
		print(f"[Monitor] Killed process group {pgid}")
	except ProcessLookupError:
		print(f"[Monitor] Process group for PID {root_pid} no longer exists.")
	except PermissionError as e:
		print(f"[Monitor] Permission denied killing PGID for {root_pid}: {e}")
	except Exception as e:
		print(f"[Monitor] Error killing PGID for {root_pid}: {e}")
		# Fallback: kill root PID only
		try:
			os.kill(root_pid, signal.SIGKILL)
			print(f"[Monitor] Killed root PID {root_pid}")
		except ProcessLookupError:
			pass
		except Exception as e2:
			print(f"[Monitor] Could not kill root PID {root_pid}: {e2}")

def monitor_and_kill_multiple(root_pid, mem_threshold_gb, timeout_sec=None, poll=0.2):
	"""
	Monitor total RSS memory usage of root process + children,
	kill process group if time/memory threshold exceeded.
	"""
	start_time = time.time()
	TIMEOUT_EVENT.clear()
	MEMORY_EVENT.clear()

	while True:
		elapsed = time.time() - start_time

		# Refresh PIDs every loop
		all_pids = get_process_tree_pids(root_pid)

		# Timeout check
		if timeout_sec is not None and elapsed >= timeout_sec and not TIMEOUT_EVENT.is_set():
			print(f"[Monitor] Time limit {timeout_sec}s reached. Killing process group of {root_pid}")
			TIMEOUT_EVENT.set()
			_kill_process_group(root_pid)
			break

		# Memory check
		total_mem_gb = 0.0
		for pid in all_pids:
			try:
				proc = psutil.Process(pid)
				total_mem_gb += proc.memory_info().rss / (1024**3)
			except (psutil.NoSuchProcess, psutil.AccessDenied):
				pass

		if total_mem_gb >= mem_threshold_gb and not MEMORY_EVENT.is_set():
			print(f"[Monitor] Memory limit exceeded ({total_mem_gb:.2f} GB >= {mem_threshold_gb:.2f} GB). Killing process group of {root_pid}")
			MEMORY_EVENT.set()
			_kill_process_group(root_pid)
			break

		# Stop if root process gone
		if not psutil.pid_exists(root_pid):
			break

		time.sleep(poll)


import subprocess
import threading
import psutil
import os
import signal
import time
import select
def run_ace_with_monitor(command, mem_limit_gb, timeout, output_path):
	"""
	Launch ACE with a watchdog that enforces memory & timeout limits.
	Detects Python MemoryError during reading and logs to Excel.
	"""
	process = subprocess.Popen(
		command,
		stdout=subprocess.PIPE,
		stderr=subprocess.PIPE,
		text=True,
		bufsize=1,
		universal_newlines=True,
		start_new_session=True
	)

	solver_pid = process.pid

	# Flags for monitor
	stop_event = threading.Event()
	memory_exceeded = threading.Event()
	timeout_exceeded = threading.Event()
	python_oom = threading.Event()  # declare before loop

	# Watchdog thread
	def watchdog():
		proc = psutil.Process(solver_pid)
		start_time = time.time()
		while not stop_event.is_set():
			if time.time() - start_time > timeout:
				timeout_exceeded.set()
				try:
					import_to_excel('stop: ACE_timeout')
					os.killpg(os.getpgid(solver_pid), signal.SIGKILL)
					print(f"[Watchdog] Timeout -> killed {solver_pid}")
				except ProcessLookupError:
					pass
				break
			try:
				mem_used_gb = proc.memory_info().rss / (1024 ** 3)
				if mem_used_gb > mem_limit_gb:
					memory_exceeded.set()
					try:
						os.killpg(os.getpgid(solver_pid), signal.SIGKILL)
						print(f"[Watchdog] Memory limit exceeded -> killed {solver_pid}")
					except ProcessLookupError:
						pass
					break
			except psutil.NoSuchProcess:
				break
			time.sleep(0.1)

	threading.Thread(target=watchdog, daemon=True).start()

	stdout_lines = []
	stderr_lines = []
	fds = [process.stdout.fileno(), process.stderr.fileno()]

	try:
		while process.poll() is None and not (memory_exceeded.is_set() or timeout_exceeded.is_set()):
			ready_fds, _, _ = select.select(fds, [], [], 0.1)
			for fd in ready_fds:
				if fd == process.stdout.fileno():
					line = process.stdout.readline()
					if line:
						stdout_lines.append(line)
				elif fd == process.stderr.fileno():
					line = process.stderr.readline()
					if line:
						stderr_lines.append(line)

		if process.poll() is None:
			os.killpg(os.getpgid(solver_pid), signal.SIGKILL)

		stdout_lines += process.stdout.read().splitlines()
		stderr_lines += process.stderr.read().splitlines()


	finally:
		process.stdout.close()
		process.stderr.close()
		stop_event.set()

	stdout_text = "\n".join(stdout_lines)
	stderr_text = "\n".join(stderr_lines)

	print(stdout_text)
	return stdout_text, stderr_text, solver_pid



if __name__ == '__main__':

	set_memory_limit_gb(MEMORY_LIMIT_GB, HARD_LIMIT_GB)

	python_pid = os.getpid()  


	
	time_used=0
	mem_used=0
	iterations=0
	ilf_empty_domain = False
	ilf_reached_timeout= False

	parser = argparse.ArgumentParser(description='Process graph matching parameters.')
	parser.add_argument('--pattern', type=str, required=True, help='Name of the pattern file (without .dot extension)')
	parser.add_argument('--target', type=str, required=True, help='Name of the target file (without .dot extension)')
	parser.add_argument('--ilf', action='store_true', help='Enable ILF option')
	parser.add_argument('--ordre', action='store_true', help='Enable order option')
	parser.add_argument('--typage', action='store_true', help='Enable typing option')
	parser.add_argument('--card', action='store_true', help='Enable typing option')
	parser.add_argument("--output", type=str, required=True) 
	parser.add_argument('--timeout', action='store_true', help='Enable timeout 120 option')
	args = parser.parse_args()

	pattern = args.pattern
	target = args.target
	ilf = args.ilf
	ordre = args.ordre
	typage = args.typage
	card= args.card
	enable_timeout= args.timeout
	output_folder= args.output
	
	# pattern  ='pan'
	# target='glc_card_01082026155531'
	
	# ilf = False
	# ordre = True
	# typage = True
	# card= True
	# enable_timeout= True
	# output_folder= "/home/etud/Bureau/projet/indicators_lines_cols/"
	
	
	target_save= target.replace('/','')

	timestamp = datetime.now().strftime("%d%m%Y%H%M%S")
	solver = ACE

	# Go to parent directory
	os.chdir(os.path.dirname(os.path.realpath(__file__)) + '/..')

	# Create a solving instance of the model
	instance = dict()
	

	Gp = nx.drawing.nx_agraph.read_dot(f'lcres/{pattern}.dot')
	Gt = nx.drawing.nx_agraph.read_dot(f'lcres/{target}.dot')
	# Gp= read_graph_from_file(f'lcres/{pattern}.txt')
	# Gt= read_graph_from_file(f'lcres/{target}.txt')
	Gpattern=Gp # save Gp before labeling it
	Gp = nx.relabel_nodes(Gp, {node: f"{node}p" for node in Gp.nodes()})
	timeout = 3600
	if ilf:
		node_dict,iterations,  time_used, mem_used = run_with_timeout(1800 , ILF, Gp, Gt) #timeout 0.5 hrs
		timeout = 1800
		if node_dict == "TIMEOUT":
			ilf_reached_timeout = True
		elif node_dict == "ERROR":
			print("ILF crashed or raised exception")
		elif node_dict is None:
			ilf_empty_domain = True

	else:
		node_dict = {}

	if ilf_reached_timeout:
		print("ilf_reached_timeout")
		import_to_excel('stop: ILF_TIMEOUT') 
	elif ilf_empty_domain:
		import_to_excel('stop: EMPTY DOMAIN ILF')
	else:

		load(instance, Gpattern, 1)
		load(instance, Gt, 2)

		print('graph loaded')
		if card: 
			card_p= cp_array(instance['card1'])
			card_t= cp_array(instance['card2'])

		preced_p_graph=instance['preced_t_1']
		preced_t_graph=instance['preced_t_2']

		

		pairs_p = [(i, elem) for i, sublist in enumerate(preced_p_graph) for elem in sublist]
		pairs_t = [(i, elem) for i, sublist in enumerate(preced_t_graph) for elem in sublist]


		# Check if required keys exist in instance before calling model
		required_keys = [
			"NV_1", "NE_1", "V1", "E1", "A1", "type1", "preced_t_1",
			"NV_2", "NE_2", "V2", "E2", "A2", "type2", "preced_t_2"
		]

		for key in required_keys:
			if key not in instance:
				print(f"Missing key in instance: {key}")


		I = model(**instance)

		

		compile()  # generates lc6.xml by default

		xml_filename = f"lc6_{pattern}_{target_save}_{timestamp}.xml"

		
		shutil.move("lc6.xml", xml_filename)

		# After ILF is done
		del Gp, Gt, node_dict, instance  # delete big objects
		gc.collect()  # force garbage collection


		ace_jar_path = "/home/etud/Bureau/projet/lib/python3.11/site-packages/pycsp3/solvers/ace/ACE-2.3.jar"



		# Where to save solver output directly
		solver_output_path = os.path.join(output_folder, "solver_output.txt")

		if os.path.exists(xml_filename):
			# command = ["java",   "-Xmx5g", "-jar", ace_jar_path, xml_filename, "-s=all", f"-t={timeout}s", "-trace", "-ev"]
			
			try:
				# ---- start ACE
				command = [
					"java",
					f"-Xmx{JAVA_HEAP_GB}g",
					"-jar", ace_jar_path,
					xml_filename,
					"-s=all",
					f"-t={timeout}s",  # optional; keep or remove; our monitor enforces hard timeout anyway
					"-trace",
					"-ev"
				]

				# Run ACE with non-blocking monitor
				stdout, stderr, solver_pid = run_ace_with_monitor(
				command=command,
				mem_limit_gb=MEMORY_LIMIT_GB,
				timeout=timeout,
				output_path=solver_output_path
			)
				solver_output = stdout or ""
				if solver_output:
						import_to_excel(stdout or "")
						print("ACE completed successfully.")

			except subprocess.TimeoutExpired:
				print(f"ACE process timed out after {timeout} seconds.")
				import_to_excel('stop: ACE_TIMEOUT')
				for pid in [python_pid, solver_pid]:
					try:
						os.kill(pid, signal.SIGKILL)
					except ProcessLookupError:
						pass

			except Exception as e:
				print(f"Error occurred: {e}")
				for pid in [python_pid]:
					try:
						os.kill(pid, signal.SIGKILL)
					except ProcessLookupError:
						pass

		else:
			print("Error: lc6.xml file not generated. Check your model definition.")
		


gc.collect()