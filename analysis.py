import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from datetime import datetime
import openpyxl
import time
import os
from openpyxl import load_workbook



def plot_charts_pos(file_folder,x_value,y_value, for_a_given_pattern='',title=''):
	# Load the Excel file
	file_path=  os.path.join(file_folder, "results.xlsx")

	df = pd.read_excel(file_path)

	df[y_value] = pd.to_numeric(df[y_value], errors='coerce')

	selected_patterns = ["pan","gap","net","bowtie","ladder"]
	if for_a_given_pattern!= '':
		df = df[df['Pattern']==for_a_given_pattern]
	else:
		df = df[df['Pattern'].isin(selected_patterns)]

	df = df[df['ILF']==False]

	# ---- Convert gt-xxxxx > gt1, gt2, gt3 ... ----
	df[x_value] = df[x_value].str.split('-').str[0].str.replace('t', 'P', regex=False).str.upper()
	df[x_value] = df[x_value] + (df.groupby(x_value).cumcount() + 1).astype(str)
	# ----------------------------------------------

	# Group by 'Pattern'  calculate average 'Effs'
	avg_effs_ilf = df.groupby([x_value])[y_value].mean().reset_index()

	# Plotting with Seaborn
	plt.figure(figsize=(12,6))
	#sns.barplot(data=avg_effs_ilf, x=x_value, y=y_value, hue=x_value,  palette='Set2')

	order = sorted(avg_effs_ilf[x_value], key=lambda x: int(''.join(filter(str.isdigit, x))))

	sns.barplot(data=avg_effs_ilf, x=x_value, y=y_value, order=order,hue=x_value,  palette='Set2')


	plt.xlabel(x_value, fontsize=20)
	plt.ylabel(y_value, fontsize=20)
	plt.title(title, fontsize=14)
	plt.xticks(rotation=45, fontsize=14)
	plt.yticks(fontsize=14)

	plt.tight_layout()
	plt.show()


def plot_effs(file_path, instance_value=0):
	df = pd.read_excel(file_path)
	instance=''
	if instance_value > 0:
		df_instance = df[df['Target_Index'] == instance_value]
		instance=f'GRC_{instance_value}'


	# Make sure 'Effs' is numeric
	df['Effs'] = pd.to_numeric(df['Effs'], errors='coerce')

	# Group by 'Pattern' and 'ILF', calculate average 'Effs'
	avg_effs_ilf = df.groupby(['Pattern','ILF'])['Effs'].mean().reset_index()

	# Plotting with Seaborn
	plt.figure(figsize=(12,6))
	sns.barplot(data=avg_effs_ilf, x='Pattern', y='Effs', hue='ILF', palette='Set2')
	plt.xlabel('Pattern',fontsize=18)
	plt.ylabel('Average Effs',fontsize=18)
	#plt.title(f'Average Effective (Effs) by Pattern with and without ILF {instance}')
	plt.xticks(rotation=45, fontsize=16)
	plt.yticks(fontsize=16)

	plt.legend(title='ILF')
	plt.tight_layout()
	plt.show()



def plot_effs_targets(file_path, instance_value=0):
	df = pd.read_excel(file_path)

	df = df[df['Pattern']=='pan']

	# Make sure 'Effs' is numeric
	df['Effs'] = pd.to_numeric(df['Effs'], errors='coerce')

	unique_targets = df['Target'].unique()
	target_mapping = {old: f'GP{i+1}' for i, old in enumerate(unique_targets)}
	df['Target_Label'] = df['Target'].map(target_mapping)

	# Group by 'Target_Label' and 'ILF', calculate average 'Effs'
	avg_effs_ilf = df.groupby(['Target_Label', 'ILF'])['Effs'].mean().reset_index()

	order = sorted(avg_effs_ilf['Target_Label'], key=lambda x: int(''.join(filter(str.isdigit, x))))


	# Plotting with Seaborn
	plt.figure(figsize=(12,6))
	sns.barplot(data=avg_effs_ilf, x='Target_Label', y='Effs', hue='ILF',order=order, palette='Set2')

	plt.xlabel('Target' ,fontsize=18)
	plt.ylabel('Average Effs', fontsize=18)
	plt.title(f'Average Effective (Effs) by Target with and without ILF')
	plt.xticks(rotation=45)
	plt.legend(title='ILF')
	plt.tight_layout()
	plt.show()





def plot_grc_data_per_target(file_path,y_data):

	df = pd.read_excel(file_path)

	df[y_data] = pd.to_numeric(df[y_data], errors='coerce')


	unique_targets = df['file4_name'].unique()
	target_mapping = {old: f'GP{i+1}' for i, old in enumerate(unique_targets)}
	df['Target_Label'] = df['file4_name'].map(target_mapping)

	# Aggregate properly (mean nodes per edge value)
	nodes_per_position_edge = (
		df.groupby('Target_Label', as_index=False)[y_data]
		.mean()
	)

	# Sort values
	nodes_per_position_edge = nodes_per_position_edge.sort_values('Target_Label')

	order = sorted(nodes_per_position_edge['Target_Label'], key=lambda x: int(''.join(filter(str.isdigit, x))))


	# Plot
	plt.figure(figsize=(12,6))
	sns.barplot(
		data=nodes_per_position_edge,
		x='Target_Label',
		y=y_data, order=order,
		palette='Set2'
	)

	plt.xlabel('Target', fontsize=18)
	plt.ylabel('#E(GRC)', fontsize=18)
	plt.xticks(rotation=45)
	plt.tight_layout()
	plt.show()




def plot_nodes_graph(file_path,x_data,y_data):

	df = pd.read_excel(file_path)

	df[y_data] = pd.to_numeric(df[y_data], errors='coerce')
	df[x_data] = pd.to_numeric(df[x_data], errors='coerce')

	# Aggregate properly (mean nodes per edge value)
	nodes_per_position_edge = (
		df.groupby(x_data, as_index=False)[y_data]
		.mean()
	)

	# Sort values
	nodes_per_position_edge = nodes_per_position_edge.sort_values(x_data)

	# Plot
	plt.figure(figsize=(12,6))
	sns.barplot(
		data=nodes_per_position_edge,
		x=x_data,
		y=y_data,
		palette='Set2'
	)

	plt.xlabel('#E(GP)', fontsize=18)
	plt.ylabel('#E(GRC)', fontsize=18)
	plt.xticks(rotation=45)
	plt.tight_layout()
	plt.show()

def compute_avg_invalid_sols(filtered: pd.DataFrame) -> dict:
	invalid_solution = filtered["Invalid solutions"].mean()
	total = len(filtered)
	return {
		"Total Instances": total,
		"Invalid sols": invalid_solution,

	}



def compute_diff(row):
	if pd.isna(row["Sols"]) or pd.isna(row["Sols_file_pos"]):
		return np.nan   # instead of "no_solution"
	else:
		return abs(row["Sols_file_pos"] - row["Sols"])



def compute_invalid_sols(file_classes,file_pos, file_linescols) :
	excel_path = os.path.join(output_folder, f"invalid_sols.xlsx")

	# === Step 1: Extract all graph names into a mapping ===
	graph_to_group = {}
	for idx, row in file_classes.iterrows():
		group = set()
		for col in file_classes.columns:
			if "name" in col:
				val = row[col]
				if pd.notna(val):
					group.add(val)
		for name in group:
			graph_to_group[name] = group

	# === Step 2: Create lookup from file_pos results ===
	file_pos_lookup = {
		(row["Pattern"], row["Target"]): row["Sols"]
		for _, row in file_pos.iterrows()
	}

	# === Step 3: For each row in file_linescols, find Sols from file_pos for group member ===
	sols_from_file_pos = []
	for idx, row in file_linescols.iterrows():
		target = row["Target"]
		pattern = row["Pattern"]
		sols_val = None
		if target in graph_to_group:
			group = graph_to_group[target]
			for member in group:
				key = (pattern, member)
				if key in file_pos_lookup:
					sols_val = file_pos_lookup[key]
					break
		sols_from_file_pos.append(sols_val)

	# === Step 4: Add results to file_linescols dataframe ===
	file_linescols["Sols_file_pos"] = sols_from_file_pos
	# === Step 5: Add difference column (Sols_file_pos - Sols) ===
	file_linescols["Sols"] = pd.to_numeric(file_linescols["Sols"], errors="coerce")
	file_linescols["Sols_file_pos"] = pd.to_numeric(file_linescols["Sols_file_pos"], errors="coerce")

	file_linescols["Invalid solutions"] = file_linescols.apply(compute_diff, axis=1)

	return file_linescols   # return enriched DataFrame, not a filename







# Helper to format mean ± std
def mean_plus_std(col_name, filtered):
	total = int(len(filtered))
	if col_name in filtered.columns and total > 0:
		mean_val = pd.to_numeric(filtered[col_name], errors="coerce").mean()
		std_val = pd.to_numeric(filtered[col_name], errors="coerce").std()
		if pd.notna(mean_val) and pd.notna(std_val):
			return f"${round(mean_val,1)} \\pm {round(std_val,1)}$"
	return "-"  # if column missing or NaN

import numpy as np
import pandas as pd
import math

def to_shared_exponent_latex(mean, std, decimals=1):

	if pd.isna(mean) or pd.isna(std):
		return np.nan

	if mean == 0:
		exponent = 0
	else:
		exponent = int(math.floor(math.log10(abs(mean))))

	scale = 10 ** exponent

	mean_scaled = round(mean / scale, decimals)
	std_scaled = round(std / scale, decimals)

	return f"$({mean_scaled} \\pm {std_scaled}) \\times 10^{{{exponent}}}$"


def compute_stats(filtered: pd.DataFrame) -> dict:
	total = int(len(filtered))
	oom_count = int((filtered["Stop"] == "Memory Limit Exceeded").sum())
	process_timeout = int((filtered["Stop"] == "process_timedout").sum())
	ilf_timeout = int((filtered["Stop"] == "ILF_TIMEOUT").sum())
	solved_instances = int((filtered["Stop"] == "FULL_EXPLORATION").sum())
	empty_domains = int((filtered["Stop"] == "EMPTY").sum())

	# Compute average Effs if column exists
	avg_effs = None
	std_effs = None
	if "Effs" in filtered.columns and total > 0:
		avg_effs = pd.to_numeric(filtered["Effs"], errors="coerce").mean()
		std_effs = pd.to_numeric(filtered["Effs"], errors="coerce").std()
		std_effs = round(std_effs, 1)
		avg_effs = round(avg_effs, 1)
	else:
		avg_effs = np.nan  
		std_effs =  np.nan

	wrgs_count = np.nan
	avg_wrgs = np.nan
	std_wrgs= None
	if "Wrgs" in filtered.columns and total > 0:
		wrgs_count = int(pd.to_numeric(filtered["Wrgs"], errors="coerce").sum()) if "Wrgs" in filtered.columns else 0
		avg_wrgs = round(pd.to_numeric(filtered["Wrgs"], errors="coerce").mean(),1) if "Wrgs" in filtered.columns and total > 0 else np.nan
		std_wrgs = round(pd.to_numeric(filtered["Wrgs"], errors="coerce").std(),1) if "Wrgs" in filtered.columns and total > 0 else np.nan
	else:
		avg_wrgs = np.nan  
		std_wrgs = np.nan
	ilf_iterations = None
	if "ilfiterations" in filtered.columns and total > 0:
		col = pd.to_numeric(filtered["ilfiterations"], errors="coerce")
		min_val = col.min()
		max_val = col.max()
		ilf_iterations = f"{round(min_val, 1)}–{round(max_val, 1)}"
	else:
		ilf_iterations = np.nan  # keep NaN if column missing or empty
  # Compute ilf time if column exists
	ilf_time = None
	if "ilftime" in filtered.columns and "Wck" in filtered.columns and total > 0:
		ilf_time = pd.to_numeric(filtered["ilftime"], errors="coerce").mean() + pd.to_numeric(filtered["Wck"], errors="coerce").mean()
		ilf_time = round(ilf_time, 1)
	else:
		ilf_time = np.nan  # keep NaN if column missing or empty

	pct = lambda n: round((n / total * 100), 1) if total > 0 else 0.0

	other = total - (oom_count + process_timeout + ilf_timeout + solved_instances + empty_domains)

	
	# Compute Wrgs and Effs as mean ± std
	# wrgs_latex = mean_plus_std("Wrgs", filtered)
	# effs_latex = mean_plus_std("Effs", filtered)
	wrgs_latex = to_shared_exponent_latex(avg_wrgs, std_wrgs)
	effs_latex = to_shared_exponent_latex(avg_effs, std_effs)

	return {
		"Total Instances": total,
		"OOM Cases": oom_count,
		"Percent OOM": pct(oom_count),
		"Empty domain cases": pct(empty_domains),
		"Process Timeout Cases": process_timeout,
		"Percent Process Timeout": pct(process_timeout),
		"ILF Timeout Cases": ilf_timeout,
		"Percent ILF Timeout": pct(ilf_timeout),
		"Solved Instances": solved_instances,
		"Percent Solved": pct(solved_instances),
		"Other/Unknown Cases": other,
		"Percent Other/Unknown": pct(other),
		"Average Effs": avg_effs,  
		"Wrgs Count": wrgs_count,
		"Wrgs (Avg)": avg_wrgs,

		"Solving time" :ilf_time,
		"Iterations": ilf_iterations,
		"Std Wrgs":std_wrgs,
		"Std Effs": std_effs,
		"wrgs_latex":wrgs_latex,
		"effs_latex":effs_latex
	}

# Function to compute results (counts + percents)
def compute_stats_pos(filtered: pd.DataFrame) -> dict:
	total = int(len(filtered))
	oom_count = int((filtered["Stop"] == "Memory Limit Exceeded").sum())
	process_timeout = int((filtered["Stop"] == "process_timedout").sum())
	ilf_timeout = int((filtered["Stop"] == "ILF_TIMEOUT").sum())
	empty_domains = int((filtered["Stop"] == "EMPTY_DOMAIN_ILF").sum())

	solved_instances = int((filtered["Stop"] == "FULL_EXPLORATION").sum())
	# Compute average Effs if column exists
	avg_effs = None
	std_effs = None
	if "Effs" in filtered.columns and total > 0:
		avg_effs = pd.to_numeric(filtered["Effs"], errors="coerce").mean()
		std_effs = pd.to_numeric(filtered["Effs"], errors="coerce").std()
		std_effs = round(std_effs, 1)
		avg_effs = round(avg_effs, 1)
	else:
		avg_effs = np.nan  # keep NaN if column missing or empty
		std_effs = np.nan
	# Compute min–max of ilf_iterations if column exists
	ilf_iterations = None
	if "ilfiterations" in filtered.columns and total > 0:
		col = pd.to_numeric(filtered["ilfiterations"], errors="coerce")
		min_val = col.min()
		max_val = col.max()
		ilf_iterations = f"{round(min_val, 1)}–{round(max_val, 1)}"
	else:
		ilf_iterations = np.nan  # keep NaN if column missing or empty
  # Compute ilf time if column exists
	ilf_time = None
	if "ilftime" in filtered.columns and "Wck" in filtered.columns and total > 0:
		ilf_time = pd.to_numeric(filtered["ilftime"], errors="coerce").mean() + pd.to_numeric(filtered["Wck"], errors="coerce").mean()
		ilf_time = round(ilf_time, 1)
	else:
		ilf_time = np.nan  # keep NaN if column missing or empty

	pct = lambda n: round((n / total * 100), 1) if total > 0 else 0.0

	other = total - (oom_count + process_timeout + ilf_timeout + solved_instances)
	avg_wrgs = None
	std_wrgs = None
	if "Wrgs" in filtered.columns and total > 0:
		avg_wrgs = pd.to_numeric(filtered["Wrgs"], errors="coerce").mean()
		std_wrgs = pd.to_numeric(filtered["Wrgs"], errors="coerce").std()
		std_wrgs= round(std_wrgs, 1)
		avg_wrgs = round(avg_wrgs, 1)
	else:
		avg_wrgs = np.nan 
		std_wrgs = np.nan 
	avg_sols = None
	std_sols = None
	if "Sols" in filtered.columns and total > 0:
		avg_sols = pd.to_numeric(filtered["Sols"], errors="coerce").mean()
		std_sols = pd.to_numeric(filtered["Sols"], errors="coerce").std()
		avg_sols = round(avg_sols, 1)
		std_sols = round(std_sols, 1)
	else:
		avg_sols = np.nan
		std_sols = np.nan


	# wrgs_latex = mean_plus_std("Wrgs", filtered)
	# effs_latex = mean_plus_std("Effs", filtered)

	wrgs_latex = to_shared_exponent_latex(avg_wrgs, std_wrgs)
	effs_latex = to_shared_exponent_latex(avg_effs, std_effs)

	return {
		"Total Instances": total,
		"OOM Cases": oom_count,
		"Percent OOM": pct(oom_count),
		"Process Timeout Cases": process_timeout,
		"Percent Process Timeout": pct(process_timeout),
		"ILF Timeout Cases": ilf_timeout,
		"Percent ILF Timeout": pct(ilf_timeout),
		"Solved Instances": solved_instances,
		"Percent Solved": pct(solved_instances),
		"Other/Unknown Cases": other,
		"Percent Other/Unknown": pct(other),
		"Average Effs": avg_effs, 
		"Wrgs":avg_wrgs,
		"Solving time" :ilf_time,
		"Iterations": ilf_iterations,
		"Empty Domain": empty_domains,
		"Avg Sols": avg_sols,
		"Std Sols": std_sols,
		"Std Wrgs":std_wrgs,
		"Std effs":std_effs,
		"wrgs_latex":wrgs_latex,
		"effs_latex":effs_latex
	}



def inner_join_models_0(df_lc, df_pos, df_generated, df_bimodel, class_name="C"):
	folder_linescols, folder_pos_path, bimodel_folder, classfile_path = get_class_files(class_name)

	# Apply filters
	mask_rc = (
		(df_lc["ILF"] == False) &
		(df_lc["Order"] == True) &
		(df_lc["Typing"] == True) &
		(df_lc["Card"] == True)
	)
	mask_pos = (df_pos["ILF"] == False)
	df_lc = df_lc[mask_rc]
	df_pos = df_pos[mask_pos]

	#fields names according to class
	if class_name == "C" or class_name == "B":
		lc_filename_field= "file3_name"
		pos_filename_field = "file4_name"
	elif class_name == "A":
		lc_filename_field= "file4_name"
		pos_filename_field = "file5_name"
	

	# Merge df_lc with df_generated on Target → file4_name
	df_lc_mapped = pd.merge(
		df_lc,
		df_generated,
		left_on="Target",
		right_on=lc_filename_field,
		how="inner"
	)

	# Merge df_pos with df_generated on Target → file3_name
	df_pos_mapped = pd.merge(
		df_pos,
		df_generated,
		left_on="Target",
		right_on=pos_filename_field,
		how="inner"
	)



	# Merge df_lc_mapped and df_pos_mapped on the generated link and Pattern
	df_lc_pos = pd.merge(
		df_lc_mapped,
		df_pos_mapped,
		on=["Pattern", pos_filename_field, lc_filename_field],
		how="inner",
		suffixes=("_lc", "_pos")
	)

	df_bimodel = df_bimodel.rename(columns={
		"Effs": "Effs_bimodel",
		"Wrgs": "Wrgs_bimodel",
		"Sols": "Sols_bimodel",
		
	})

	df_joined = pd.merge(
		df_lc_pos,
		df_bimodel,
		left_on=[lc_filename_field, "Pattern"],
		right_on=["Target", "Pattern"],
		how="inner",   
		suffixes=("", "_bimodel")
	)

	# Keep relevant columns and rename for clarity
	df_joined= df_joined.rename(
		columns={
			"Target_lc": "Target_LC",
			"Target_pos": "Target_POS",
			"Effs_pos":"Effs_pos",
			"Effs_lc":"Effs_lc",
			"Wrgs_pos":"Wrgs_pos",
			"Wrgs_lc":"Wrgs_lc",
			"Sols_lc":"Sols_lc",
			"Sols_pos": "Sols_pos",
			"Effs_bimodel":"Effs_bimodel",
			"Wrgs_bimodel":"Wrgs_bimodel",
			"Sols_bimodel":"Sols_bimodel",
			 "Stop": "Stop_bimodel" 
		}
	)


	df_final = df_joined[["Pattern", "Target_LC", "Target_POS", "Effs_pos", "Effs_lc","Effs_bimodel", "Wrgs_pos", "Wrgs_lc","Wrgs_bimodel" ,"Sols_lc", "Sols_pos", "Sols_bimodel", "Stop_bimodel"]]
	excel_path = os.path.join(folder_pos_path, f"join_results.xlsx")

	df_final.to_excel(excel_path, index=False)
	print(f"Pos and rows/cols results saved to {excel_path}")
	return df_final






def inner_join_models(df_lc, df_pos, df_generated, df_bimodel,df_bimodel_ilf, class_name="C"):
	folder_linescols, folder_pos_path, bimodel_folder,bimodel_ilf, classfile_path = get_class_files(class_name)

	# Apply filters
	mask_rc = (
		(df_lc["ILF"] == False) &
		(df_lc["Order"] == True) &
		(df_lc["Typing"] == True) &
		(df_lc["Card"] == True)
	)
	mask_pos = (df_pos["ILF"] == False)
	df_lc = df_lc[mask_rc]
	df_pos = df_pos[mask_pos]

	#fields names according to class
	if class_name == "C" or class_name == "B":
		lc_filename_field= "file3_name"
		pos_filename_field = "file4_name"
	elif class_name == "A":
		lc_filename_field= "file4_name"
		pos_filename_field = "file5_name"
	

	# Merge df_lc with df_generated on Target → file4_name
	df_lc_mapped = pd.merge(
		df_lc,
		df_generated,
		left_on="Target",
		right_on=lc_filename_field,
		how="inner"
	)

	# Merge df_pos with df_generated on Target → file3_name
	df_pos_mapped = pd.merge(
		df_pos,
		df_generated,
		left_on="Target",
		right_on=pos_filename_field,
		how="inner"
	)



	# Merge df_lc_mapped and df_pos_mapped on the generated link and Pattern
	df_lc_pos = pd.merge(
		df_lc_mapped,
		df_pos_mapped,
		on=["Pattern", pos_filename_field, lc_filename_field],
		how="inner",
		suffixes=("_lc", "_pos")
	)

	df_bimodel = df_bimodel.rename(columns={
		"Effs": "Effs_bimodel",
		"Wrgs": "Wrgs_bimodel",
		"Sols": "Sols_bimodel",
		
	})


	df_bimodel_ilf = df_bimodel_ilf.rename(columns={
		"Effs": "Effs_bimodel_ilf",
		"Wrgs": "Wrgs_bimodel_ilf",
		"Sols": "Sols_bimodel_ilf",
		"Stop": "Stop_bimodel_ilf"
	})
	df_joined = pd.merge(
		df_lc_pos,
		df_bimodel,
		left_on=[lc_filename_field, "Pattern"],
		right_on=["Target", "Pattern"],
		how="inner",   
		suffixes=("", "_bimodel")
	)

	df_joined = pd.merge(
	df_joined,
	df_bimodel_ilf,
	left_on=[lc_filename_field, "Pattern"],
	right_on=["Target", "Pattern"],
	how="inner",
	suffixes=("", "_ilf")
	)


	# Keep relevant columns and rename for clarity
	df_joined= df_joined.rename(
		columns={
			"Target_lc": "Target_LC",
			"Target_pos": "Target_POS",
			"Effs_pos":"Effs_pos",
			"Effs_lc":"Effs_lc",
			"Wrgs_pos":"Wrgs_pos",
			"Wrgs_lc":"Wrgs_lc",
			"Sols_lc":"Sols_lc",
			"Sols_pos": "Sols_pos",
			"Effs_bimodel":"Effs_bimodel",
			"Wrgs_bimodel":"Wrgs_bimodel",
			"Sols_bimodel":"Sols_bimodel",
			 "Stop": "Stop_bimodel" 
		}
	)


	df_final = df_joined[[
		"Pattern", "Target_LC", "Target_POS",
	"Effs_pos", "Effs_lc", "Effs_bimodel", "Effs_bimodel_ilf",
	"Wrgs_pos", "Wrgs_lc", "Wrgs_bimodel", "Wrgs_bimodel_ilf",
	"Sols_lc", "Sols_pos", "Sols_bimodel", ]]

	excel_path = os.path.join(folder_pos_path, f"join_results.xlsx")

	df_final.to_excel(excel_path, index=False)
	print(f"Pos and rows/cols results saved to {excel_path}")
	return df_final


def import_excel_sols(df, filters):
	
	# Normalize the Stop column just in case of extra spaces / case differences
	if "Stop" in df.columns:
		df["Stop"] = df["Stop"].astype(str).str.strip()
	# Helper to convert booleans to VRAI/FAUX for the output table
	vf = {True: "VRAI", False: "FAUX"}

	rows = []
	for name, cond in filters:
		mask = (
			(df["ILF"] == cond["ILF"]) &
			(df["Order"] == cond["Order"]) &
			(df["Typing"] == cond["Typing"]) &
			(df["Card"] == cond["Card"]) 
		)
		stats = compute_avg_invalid_sols(df[mask])
		

		row = {
			"Filter": name,
			"ILF": vf[cond["ILF"]],
			"Order": vf[cond["Order"]],
			"Typing": vf[cond["Typing"]],
			"Card": vf[cond["Card"]],
			**stats,
		}
		rows.append(row)

	# Explicit column order to avoid any misalignment in Excel
	columns = [
	"Filter", "ILF", "Order", "Typing", "Card",
	"Total Instances", "Invalid sols"
	]
	excel_path = os.path.join(folder_linescols, f"avg_sols.xlsx")


	results_df = pd.DataFrame(rows, columns=columns)

	# Write to Excel cleanly
	with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
		results_df.to_excel(writer, index=False)

	print(f"Avg sols saved to {excel_path}")



def import_excel_stats(df, filters):
	# Normalize the Stop column
	if "Stop" in df.columns:
		df["Stop"] = df["Stop"].astype(str).str.strip()

	# Convert booleans to VRAI/FAUX
	vf = {True: "VRAI", False: "FAUX"}

	rows = []
	for name, cond in filters:
		mask = (
			(df["ILF"] == cond["ILF"]) &
			(df["Order"] == cond["Order"]) &
			(df["Typing"] == cond["Typing"]) &
			(df["Card"] == cond["Card"])
		)
		stats = compute_stats(df[mask])

		row = {
			"Filter": name,
			"ILF": vf[cond["ILF"]],
			"Order": vf[cond["Order"]],
			"Typing": vf[cond["Typing"]],
			"Card": vf[cond["Card"]],
			**stats,
		}
		rows.append(row)

	# Column order
	columns = [
		"Filter", "ILF", "Order", "Typing", "Card",
		"Total Instances",
		"Percent OOM",
		 "Percent Solved",
		"Average Effs","Std Effs", "Solving time", "Iterations","Wrgs (Avg)","Std Wrgs", "wrgs_latex", "effs_latex"
	]

	results_df = pd.DataFrame(rows, columns=columns)
	excel_path = os.path.join(folder_linescols, "avg_sols.xlsx")

	# Write DataFrame to Excel
	results_df.to_excel(excel_path, index=False)

	# Open workbook to add JOINDRE.TEXTE formula
	wb = load_workbook(excel_path)
	ws = wb.active

	# Determine the last column and last row
	last_col_letter = ws.cell(row=1, column=len(columns)).column_letter
	first_data_row = 2
	last_data_row = ws.max_row

	# Add formula in the row below the last data row
	formula_row = last_data_row + 1
	formula = f'=JOINDRE.TEXTE("&";VRAI;{last_col_letter}{first_data_row}:{last_col_letter}{last_data_row})'.upper()
	ws[f"{last_col_letter}{formula_row}"] = formula

	wb.save(excel_path)
	print(f"Aligned results saved to {excel_path}")



def import_excel_stats_pos(df, filters):
	# Normalize the Stop column
	if "Stop" in df.columns:
		df["Stop"] = df["Stop"].astype(str).str.strip()

	# Convert booleans to VRAI/FAUX
	vf = {True: "VRAI", False: "FAUX"}

	rows = []
	for name, cond in filters:
		mask = (
			(df["ILF"] == cond["ILF"]) 
		)
		stats = compute_stats_pos(df[mask])

		row = {
			"Filter": name,
			"ILF": vf[cond["ILF"]],

			**stats,
		}
		rows.append(row)

	# Column order
	columns = [
		"Filter", "ILF", "Order", "Typing", "Card",
		"Total Instances", "Percent OOM","Solved Instances", 
		"Average Effs", "Solving time", "Iterations", "Wrgs","Std Wrgs", "Avg Sols", "Std Sols","Std effs"
		,"wrgs_latex","effs_latex"
	]

	results_df = pd.DataFrame(rows, columns=columns)
	excel_path = os.path.join(folder_pos_path, "avg_sols.xlsx")

	# Write DataFrame to Excel
	results_df.to_excel(excel_path, index=False)

	# Open workbook to add JOINDRE.TEXTE formula
	wb = load_workbook(excel_path)
	ws = wb.active

	# Determine the last column and last row
	last_col_letter = ws.cell(row=1, column=len(columns)).column_letter
	first_data_row = 2
	last_data_row = ws.max_row

	wb.save(excel_path)
	print(f"Aligned results saved to {excel_path}")



def import_stats_pattern_pos(df, filters):
	# Normalize the Stop column
	if "Stop" in df.columns:
		df["Stop"] = df["Stop"].astype(str).str.strip()

	# Convert booleans to VRAI/FAUX
	vf = {True: "VRAI", False: "FAUX"}

	rows = []
	for name, cond in filters:
		mask = pd.Series(True, index=df.index)
		for col, value in cond.items():
			mask &= (df[col] == value)

		stats = compute_stats_pos(df[mask])

		row = {
			"Filter": name,
			"ILF": vf[cond["ILF"]],
			"Pattern":cond["Pattern"],
			**stats,
		}
		rows.append(row)

	# Column order
	columns = [
		"Filter", "ILF","Pattern",
		"Total Instances",
		"Solved Instances", "Percent Solved",
		"Average Effs", "Solving time", "Iterations","Wrgs","Std Wrgs", "Avg Sols", "Std Sols","Std effs","wrgs_latex","effs_latex"
	]

	results_df = pd.DataFrame(rows, columns=columns)
	excel_path = os.path.join(folder_pos_path, "sols_per_pattern.xlsx")

	# Write DataFrame to Excel
	results_df.to_excel(excel_path, index=False)

	# Open workbook to add JOINDRE.TEXTE formula
	wb = load_workbook(excel_path)
	ws = wb.active

	# Determine the last column and last row
	last_col_letter = ws.cell(row=1, column=len(columns)).column_letter
	first_data_row = 2
	last_data_row = ws.max_row

	# Add formula in the row below the last data row
	formula_row = last_data_row + 1
	formula = f'=JOINDRE.TEXTE("&";VRAI;{last_col_letter}{first_data_row}:{last_col_letter}{last_data_row})'.upper()
	ws[f"{last_col_letter}{formula_row}"] = formula

	wb.save(excel_path)
	print(f"Results per pattern saved to {excel_path}")





def import_stats_pattern_lc(df, filters):
	# Normalize the Stop column
	if "Stop" in df.columns:
		df["Stop"] = df["Stop"].astype(str).str.strip()

	# Convert booleans to VRAI/FAUX
	vf = {True: "VRAI", False: "FAUX"}

	rows = []
	for name, cond in filters:
		mask = pd.Series(True, index=df.index)
		for col, value in cond.items():
			mask &= (df[col] == value)

		stats = compute_stats(df[mask])

		row = {
			"Filter": name,
			"ILF": vf[cond["ILF"]],
			"Order": vf[cond["Order"]],
			"Typing": vf[cond["Typing"]],
			"Card": vf[cond["Card"]],
			"Pattern": cond["Pattern"],
			**stats,
		}
		rows.append(row)

	columns = [
		"Filter", "ILF", "Order", "Typing", "Card","Pattern","Percent Solved",
		"Average Effs", "Solving time", "Iterations","Wrgs (Avg)", "Std Effs", "Std Wrgs", "wrgs_latex", "effs_latex"
	]


	results_df = pd.DataFrame(rows, columns=columns)
	excel_path = os.path.join(folder_linescols, "sols_per_pattern.xlsx")

	# Write DataFrame to Excel
	results_df.to_excel(excel_path, index=False)

	# Open workbook to add JOINDRE.TEXTE formula
	wb = load_workbook(excel_path)
	ws = wb.active

	# Determine the last column and last row
	last_col_letter = ws.cell(row=1, column=len(columns)).column_letter
	first_data_row = 2
	last_data_row = ws.max_row

	wb.save(excel_path)
	print(f"Results per pattern saved to {excel_path}")


def get_class_files(class_name):
	folder_linescols= ""
	folder_pos_path = ""
	classfile_path = ""
	bimodel_folder=""
	bimodel_ilf_folder =""
	if class_name == 'A':
		folder_linescols= "/home/etud/Bureau/projet/indicators_lines_cols/run_class_b_15082025152720"
		folder_pos_path = "/home/etud/Bureau/projet/indicators_pos/run_10092025142625"
		classfile_path = "/home/etud/Bureau/projet/fichiers/csp/generated_files_db/class_b_28072025080534.xlsx"
		bimodel_folder="/home/etud/Bureau/projet/indicators_pos/run_12032026144957/"
		bimodel_ilf_folder="/home/etud/Bureau/projet/indicators_pos/run_23032026141936/"
	elif class_name == 'B':
		folder_linescols= "/home/etud/Bureau/projet/indicators_lines_cols/run_class_200x50_24092025111514"
		folder_pos_path = "/home/etud/Bureau/projet/indicators_pos/run_23092025154227"
		classfile_path = "/home/etud/Bureau/projet/fichiers/csp/generated_files_db/file_22092025154919.xlsx"
	elif class_name == 'C':
		folder_linescols= "/home/etud/Bureau/projet/indicators_lines_cols/run_class_200x50_28112025093939"
		folder_pos_path = "/home/etud/Bureau/projet/indicators_pos/run_27112025103942"
		classfile_path = "/home/etud/Bureau/projet/fichiers/csp/generated_files_db/file_25112025133728_copy_plot.xlsx"
	elif class_name == 'X':
		folder_linescols= "/home/etud/Bureau/projet/indicators_lines_cols/run_05032026104355"
		folder_pos_path = "/home/etud/Bureau/projet/indicators_pos/run_27112025103942"
		classfile_path = "/home/etud/Bureau/projet/fichiers/csp/generated_files_db/file_25112025133728.xlsx"

	return folder_linescols, folder_pos_path, classfile_path,bimodel_folder, bimodel_ilf_folder


#remember to run merge4 and 3 before
if __name__ == '__main__':
	class_name='A'
	folder_linescols, folder_pos_path, classfile_path, bimodel_folder, bimodel_ilf_folder = get_class_files(class_name)

	timestamp = datetime.now().strftime("%d%m%Y%H%M%S")

	output_folder = "/home/etud/Bureau/projet/indicators_lines_cols/"
	file_linescols_path=  os.path.join(folder_linescols, "results.xlsx")
	file_pos_path =os.path.join(folder_pos_path, "results.xlsx")
	file_bimodel =os.path.join(bimodel_folder,"results.xlsx")
	file_bimodel_ilf =os.path.join(bimodel_ilf_folder,"results.xlsx")

	file_classes = pd.read_excel(classfile_path)
	file_pos = pd.read_excel(file_pos_path)
	file_linescols = pd.read_excel(file_linescols_path)
	df_generated= pd.read_excel(classfile_path)
	df_bimodel=pd.read_excel(file_bimodel)
	df_bimodel_ilf=pd.read_excel(file_bimodel_ilf)



	patterns = [ "pan", "galaxy"]
	ilf_values = [True, False]
	filters_patterns_pos = []

	for pattern in patterns:
		for ilf in ilf_values:
			name = f"{pattern} – ILF {ilf}"
			filters_patterns_pos.append(
				(name, {"ILF": ilf, "Pattern": pattern})
			)



	filters = [
		("Filter_1", {"ILF": False, "Order": False, "Typing": False, "Card": False}),
		("Filter_2", {"ILF": True, "Order": True,  "Typing": True,  "Card": True}),
		("Filter_3", {"ILF": True,  "Order": True,  "Typing": True,  "Card": False}),
		("Filter_4", {"ILF": False,  "Order": True,  "Typing": True,  "Card": False}),
		("Filter_5", {"ILF": False, "Order": True,  "Typing": True,  "Card": True}),

	]


	filters_pattern_lc = [
	(f"{name}_{pattern}", {**conds, "Pattern": pattern})
	for name, conds in filters
	for pattern in patterns
	]
	dflc= pd.read_excel(file_linescols_path)



	dfpos= pd.read_excel(file_pos_path)
	filters_pos = [
		("Filter_1", {"ILF": False}),
		("Filter_2", {"ILF": True,})]

	#plot_effs(file_pos_path,0)
	#plot_charts_pos(folder_pos_path,'Target','Sols','pan')
	#plot_nodes_graph(classfile_path,'file4_edges','file2_edges')
	#plot_grc_data_per_target(classfile_path,'file2_edges')
	# import_excel_stats_pos(dfpos, filters_pos)
	# import_excel_stats(dflc, filters)
	# import_stats_pattern_pos(dfpos, filters_patterns_pos)
	#import_stats_pattern_lc(dflc, filters_pattern_lc)
	df_lc_pos=inner_join_models(dflc,dfpos, df_generated,df_bimodel,df_bimodel_ilf,class_name)