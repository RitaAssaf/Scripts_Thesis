from enum import Enum, auto
from pycsp3 import *
import subprocess
import os
import sys
import argparse
import pygraphviz as pgv
from pathlib import Path
import inflect
import networkx as nx
import pydot
import re
import openpyxl
import subprocess
import psutil
import threading
import csv
import time
import shutil
from typing import List
from collections import Counter
from class_label_order import LabelOrder

from datetime import datetime 
from ilf import run_with_timeout,get_label,hopcroft, build_partial_order_3, build_partial_order_4,hopcroft_multiset,dir_ILF, ILF
from itertools import combinations
from datetime import datetime



# def import_log_file(message: str, log_file: str = "app1.log"):
# 	with open(log_file, "a") as f:
# 		f.write(f"{datetime.now():%Y-%m-%d %H:%M:%S} - {message}\n")



def import_log_file(message: str):
	# Create a unique log file name based on current date and time
	timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
	log_file = f"app_log_{timestamp}.log"
	
	# Ensure logs are stored in a 'logs' directory (optional but recommended)
	os.makedirs("logs", exist_ok=True)
	log_path = os.path.join("logs", log_file)

	# Write the log message
	with open(log_path, "a") as f:
		f.write(f"{datetime.now():%Y-%m-%d %H:%M:%S} - {message}\n")

	print(f"Log written to: {log_path}")


class Angle(str, Enum):
	Horizontal = 'Horizontal'
	Vertical = 'Vertical'


def get_node_label(G, label):
	#return [n for n, d in G.nodes(data=True) if d.get("label") == "Person"][0]
	return {n.get_name(): n for n in G.get_nodes()}[label] #graphiz file


def process_mapping( instance, mapping_n, mapping, target,pattern):

	print(f'{mapping_n}.\t[{", ".join(instance["V1"])}] -> [{", ".join(list(map(lambda i: instance["realnames2"][i], mapping)))}]')

	# Plot mapping
	graph = pydot.graph_from_dot_file(f'lcres/{target}.dot')[0]
	for v1 in range(instance['NV_1']):#
		# prend le label du noeud target qui est dans l'array V2 a l'indice solution trouvé par le solveur
		#donc le solveur retourne array des indices des noeuds de la solution 
		node=instance['V2'][mapping[v1]] 
		get_node_label(graph, node).set_label(instance['V1'][v1])
		#graph.get_node(node)[0].set_label(instance['V1'][v1])
		get_node_label(graph, node).set_style('filled')
		graph.write_png(f'res/{pattern}_in_{target}_{mapping_n}.png')


# Angle of the edge of a graph
class Angle(str, Enum):
	Horizontal = 'Horizontal'
	Vertical = 'Vertical'

def domain_n(i, NV_2):
	
	indices_l= [i for i, type in enumerate(instance['lc_type_target'] ) if type == 'l']

	indices_c= [i for i, type in enumerate(instance['lc_type_target'] ) if type == 'c']
	#si le type du noeud patterne est l donc il aura les indices(qui sont les noms aussi) des noeuds l
	
	if not ilf and not typage:	
		return range(NV_2 )
	elif not ilf and typage:	
		if instance['lc_type_pattern'][i]== 'l':
			return indices_l
		else: return indices_c
	elif ilf and not typage:	
		return node_dict[f'{i}p']
	elif ilf and typage:
		ilf_domain= node_dict[f'{i}p']
		if instance['lc_type_pattern'][i]== 'l':
			return list(set(ilf_domain) & set(indices_l))
		else: return list(set(ilf_domain) & set(indices_c))
	

def dom_preced(i: int, j: int, target):
	if target:
		if 0 <= i < len(preced_t_graph) and 0 <= j < len(preced_t_graph[i]):
			return [preced_t_graph[i][j]] 
		return [-1] 
	else:
		if 0 <= i < len(preced_p_graph) and 0 <= j < len(preced_p_graph[i]):
			return [preced_p_graph[i][j]] 
		return [-1] 

def fill_matrix(lst_of_lsts,length):
	
	result = [sublist + [-1] * (length - len(sublist)) for sublist in lst_of_lsts]  # Fill shorter lists with -1
	
	return result


def import_to_excel(solver_output):
	#Step 1: Clean ANSI escape sequences
	ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
	clean_output = ansi_escape.sub('', solver_output)

	# Step 2: Extract metrics with individual fallback regexes
	effs_matches = re.findall(r'effs\s*:\s*(\d+)', clean_output)
	effs = effs_matches[-1] if effs_matches else "-"

	stop_match = re.search(r'stop\s*:\s*([A-Z_]+)', clean_output)
	stop = stop_match.group(1) if stop_match else "-"

	wck_matches = re.findall(r'wck\s*:\s*(\d+\.\d+)', clean_output)
	wck = wck_matches[-1] if wck_matches else "-"

	cpu_match = re.search(r'cpu\s*:\s*(\d+\.\d+)', clean_output)
	cpu = cpu_match.group(1) if cpu_match else "-"

	mem_match = re.search(r'mem\s*:\s*(\S+)', clean_output)
	mem = mem_match.group(1) if mem_match else "-"

	unsat_match = re.search(r'\bs\s+UNSATISFIABLE\b', clean_output)
	unsat = "Yes" if unsat_match else "No"

	wrong_dec_match = re.search(r'd\s+WRONG\s+DECISIONS\s+(\d+)', clean_output)
	wrong_dec = wrong_dec_match.group(1) if wrong_dec_match else "-"

	found_sols_match = re.search(r'd\s+FOUND\s+SOLUTIONS\s+(\d+)', clean_output)
	found_sols = found_sols_match.group(1) if found_sols_match else "-"

	complete_match = re.search(r'd\s+COMPLETE\s+EXPLORATION', clean_output)
	complete = "Yes" if complete_match else "No"

	# Step 3: Create Excel workbook
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Indicators"

	# Step 3: Create Excel workbook
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Indicators"

	headers = [
		"Pattern", "Target", "ILF", "Order", "Typing", "Card", "Run",
		"Dpts", "Effs", "Fails", "Wrgs", "Wck", "Ngds", "Sols",
		"Stop", "CPU", "Mem", "UNSAT", "WrongDec", "FoundSols", "Complete",
		"ilftime", "ilfmemo","ilfiterations"
	]
	ws.append(headers)

	# Step 4: Prepare and append row
	run_count = 1
	row = [
		pattern, target, ilf, ordre, typage, card,
		f"run{run_count}",  # Run ID
		"-",                # Dpts (not extracted by fallback)
		effs,
		"-",                # Fails (not extracted by fallback)
		wrong_dec,          # Wrgs (using wrong decisions)
		wck,
		"-",                # Ngds (not extracted by fallback)
		found_sols,
		stop,
		cpu,
		mem,
		unsat,
		wrong_dec,
		found_sols,
		complete,
		0,
		0,
		0
	]
	ws.append(row)

	# Step 5: Save Excel file with memory error handling
	timestamp = datetime.now().strftime("%d%m%Y%H%M%S")
	file_name = os.path.join(output_folder, f"indicators_{pattern}_{target}_{timestamp}.xlsx")

	try:
		wb.save(file_name)
		print(f"Data has been written to {file_name}")

	except MemoryError as mem_err:
		print(f"MemoryError when saving Excel: {mem_err}")
		# Free as much memory as possible
		import gc
		gc.collect()

		# Try saving a minimal Excel file with just an error message
		try:
			wb_minimal = openpyxl.Workbook()
			ws_min = wb_minimal.active
			ws_min.title = "Indicators"
			ws_min.append(["Status"])
			ws_min.append(["stop: ACE_MEM_LIMIT"])
			min_file_name = os.path.join(output_folder, f"indicators_{pattern}_{target}_{timestamp}.xlsx")
			wb_minimal.save(min_file_name)
			print(f"Minimal memory error report saved to {min_file_name}")
		except Exception as e2:
			print(f"Failed to save minimal Excel report: {e2}")

	except Exception as e:
		print(f"Error saving excel file: {e}")
		sys.exit(1)




def load(instance, graph_file, id, node_dict={}):

	# Load graph
	graph = nx.drawing.nx_agraph.read_dot(graph_file)

	edges = graph.edges  # if multiedgeview add strict to graphiz
	lenedges = len(edges)
	labels = nx.get_node_attributes(graph, 'label')
	types = nx.get_node_attributes(graph, 'type')
	preceds = nx.get_node_attributes(graph, 'preceds')  # Extract preceds attribute
	cards = nx.get_node_attributes(graph, 'card')
	realnames= nx.get_node_attributes(graph, 'realname')
	graphs = pydot.graph_from_dot_file(graph_file)
	pydot_graph = graphs[0]

	# Create a mapping of node labels to indices
	label_to_index = {label: int(node) for node, label in labels.items()}

	# Build preced_t list of lists based on preceds attribute
	preced_t = []
	for i in range(len(labels)):
		node_name = str(i)
		if node_name in preceds:
			# Convert preceds node labels to their corresponding indices
			preced_indices = [label_to_index[succ] for succ in preceds[node_name].split(',') if succ]
			preced_t.append(preced_indices)
		else:
			preced_t.append([])  # No preceds for this node


	instance[f'lc_preced_{id}'] = preced_t  # Save to instance for later use

	# Set instance data
	instance[f'lc_NV_{id}'] = graph.number_of_nodes()
	instance[f'lc_NE_{id}'] = graph.number_of_edges()
	instance[f'lc_V_{id}'] = [name for name in labels.values()]
	instance[f'lc_type_{id}'] = [type for type in types.values()]
	instance[f'lc_E_{id}'] = list()

	instance[f'lc_card_{id}'] = [int(c) for c in cards.values()]
	for n1, n2 in edges:
		instance[f'lc_E_{id}'].append((int(n1), int(n2)))
	instance[f'lc_A_{id}'] = list()
	instance[f'lc_realnames_{id}'] = [name for name in realnames.values()]
	breakpoint=1

def load_position(instance, graph_file, id, node_dict={}):
	graph = nx.drawing.nx_agraph.read_dot(graph_file)

	edges = [(u, v, d['angle']) for u, v, d in graph.edges(data=True) if 'angle' in d]
	labels = nx.get_node_attributes(graph, 'label')
	angles = nx.get_edge_attributes(graph, 'angle')
	graphs = pydot.graph_from_dot_file(graph_file)
	pydot_graph = graphs[0]
	nodes = pydot_graph.get_nodes()
	#print("Total nodes in PyDot:", len(nodes))
	#print("Node labels:", labels)

	# Set instance data
	instance[f'pos_NV_{id}'] = graph.number_of_nodes()
	instance[f'pos_NE_{id}'] = graph.number_of_edges()
	instance[f'pos_V_{id}'] = [label for label in labels.values()]
	instance[f'pos_E_{id}'] = list()
	instance[f'pos_A_{id}'] = list()

	for n1, n2, a in edges:
		instance[f'pos_E_{id}'].append((int(n1), int(n2)))
		key = (n1, n2, 0) if (n1, n2, 0) in angles else (n1, n2)
		instance[f'pos_A_{id}'].append('Horizontal' if angles[key] == '0' else 'Vertical')

def contains(lc_index: int, token: str) -> bool:
	realnames=list(instance['lc_realnames_target']) 
	s= realnames[lc_index]
	if s.startswith('l') or s.startswith('c'):
		s = s[1:]
	tokens = list(map(int,s.split('.')))
	return token in tokens

def tokens(lc_index: int) -> List[int]:
	realnames=list(instance['lc_realnames_pattern']) 
	s= realnames[lc_index]
	if s.startswith('l') or s.startswith('c'):
		s = s[1:]
	tokens = list(map(int,s.split('.')))
	return  tokens


#Function model takes as parameter nodes,edges and retrun var array
def model(

	# =============== DATA

	# Pattern graph G1 = (V1, E1), with |V1| = lc_NV_1, |E1| = lc_NV_1, and A1 the angle of each edges in E1
	lc_NV_pattern: int,
	lc_NE_pattern: int,
	lc_V_pattern: list[str],
	lc_E_pattern : list[list[int]],
	lc_A_pattern: list[Angle],
	lc_type_pattern:list[str],
	lc_preced_pattern:list[list[int]],

	# Target graph G2 = (V2, E2), with |V2| = lc_NV_2, |E2| = lc_NV_2, and A2 the angle of each edges in E2
	lc_NV_target: int,
	lc_NE_target: int,
	lc_V_target: list[str],
	lc_E_target: list[list[int]],
	lc_A_target: list[Angle],
	lc_type_target:list[str],
	lc_preced_target:list[list[int]],
	lc_card_pattern=[],
	lc_card_target=[],
	lc_realnames_pattern=[],
	lc_realnames_target=[],

	# Position pattern graph G1 = (V1, E1), with |V1| = lc_NV_1, |E1| = lc_NV_1, and A1 the angle of each edges in E1

	pos_NV_pattern: int=0,
	pos_NE_pattern: int=0,
	pos_V_pattern: list[str]=[],
	pos_E_pattern: list[list[int]]=[],
	pos_A_pattern: list[Angle]=[],

	# Target graph G2 = (V2, E2), with |V2| = lc_NV_2, |E2| = lc_NV_2, and A2 the angle of each edges in E2
	pos_NV_target: int=0,
	pos_NE_target: int=0,
	pos_V_target: list[str]=[],
	pos_E_target: list[list[int]]=[],
	pos_A_target: list[Angle]=[]
	):
		# =============== VARIABLE

		# ensembleLC = {0,1}
		# ensembleT={ 5,20,200}
		ensemble_LC_target = set(range(lc_NV_target))
		ensemble_LC_pattern = set(range(lc_NV_pattern))
		
		ensemble_tokens_target=set(range(pos_NV_target))

		map_pos = VarArray(size=pos_NV_pattern, dom=lambda i: range(pos_NV_target))

		map_lc = VarArray(size=lc_NV_pattern, dom=lambda i: domain_n(i, lc_NV_target))

		if card:
			cardinality_p= VarArray(size= lc_NV_pattern, dom=range(15))
			cardinality_t=VarArray(size= lc_NV_target, dom=range(15))

		T = {( i , j ) for i , j in lc_E_target } | {( j , i ) for i , j in lc_E_target }

		# =============== CONSTRAINT
		satisfy(

			AllDifferent(map_lc),

			[( map_lc[ i ] , map_lc[ j ]) in T for (i , j ) in lc_E_pattern ] ,

		)

		if ordre :
			satisfy (
				[( map_lc[ i ] , map_lc[ j ]) in pairs_t for (i , j ) in pairs_p ]
			)

		if card :
			satisfy (
				[cardinality_p[n] == card_p[n] for n in range(lc_NV_pattern)],
				[cardinality_t[n] == card_t[n] for n in range(lc_NV_target)],
				[cardinality_t[map_lc[i]] == cardinality_p[i] for i in range(lc_NV_pattern)]
			)

	#########Position model#################################################################################################################
		satisfy (
			AllDifferent(map_pos),

			[
			Exist(
			(map_pos[pos_E_pattern[e1][0]] == pos_E_target[e2][0]) & (map_pos[pos_E_pattern[e1][1]] == pos_E_target[e2][1])
			for e2 in range(pos_NE_target) if pos_A_pattern[e1]==pos_A_target[e2]
			) for e1 in range(pos_NE_pattern)
			])

		satisfy (
		[
		If (map_lc[line_pattern].among(ensemble_LC_target - {l for l in ensemble_LC_target if contains(l,t)})
		,
		Then =map_pos[i].among(ensemble_tokens_target- {t}) 
		)for line_pattern in ensemble_LC_pattern for i in tokens(line_pattern) for t in ensemble_tokens_target 
		],
		)



		# for i in range(lc_NV_pattern):
		# 	print(map_lc[i].dom)
		# for i in range(pos_NV_pattern):
		# 	print(map_pos[i].dom)
		#print(posted())
		return map_lc,map_pos



if __name__ == '__main__':

	#ilf_reached_timeout= False
	# pattern='pan'
	# target='glc_card_09222025154928'
	# pos_pattern='pan'
	# pos_target='gt-09222025154928'

	# pattern='pan'
	# target='panp1_lctr-06082025165701'
	# pos_pattern='pan'
	# pos_target='gt-06082025165701'
	
	pattern='pan'
	target='glc_card_07282025080536'
	pos_pattern='pan'
	pos_target='gt-07282025080536'
	
	
	ilf = False
	ordre = True
	typage = True
	card= True
	enable_timeout= True
	output_folder= "/home/etud/Bureau/projet/indicators_lines_cols/"

	# parser = argparse.ArgumentParser(description='Process graph matching parameters.')
	# parser.add_argument('--pattern', type=str, required=True, help='Name of the pattern file (without .dot extension)')
	# parser.add_argument('--target', type=str, required=True, help='Name of the target file (without .dot extension)')
	# parser.add_argument('--pos_target', type=str, required=True, help='Name of the pattern file (without .dot extension)')
	# parser.add_argument('--pos_pattern', type=str, required=True, help='Name of the target file (without .dot extension)')
	
	# parser.add_argument('--ilf', action='store_true', help='Enable ILF option')
	# parser.add_argument('--ordre', action='store_true', help='Enable order option')
	# parser.add_argument('--typage', action='store_true', help='Enable typing option')
	# parser.add_argument('--card', action='store_true', help='Enable typing option')
	# parser.add_argument("--output", type=str, required=True) 
	# parser.add_argument('--timeout', action='store_true', help='Enable timeout 120 option')
	# args = parser.parse_args()

	# pos_target=args.pos_target
	# pos_pattern=args.pos_pattern
	# pattern = args.pattern
	# target = args.target
	# ilf = args.ilf
	# ordre = args.ordre
	# typage = args.typage
	# card= args.card
	# enable_timeout= args.timeout
	# output_folder= args.output


	timestamp = datetime.now().strftime("%d%m%Y%H%M%S")

	solver = ACE
	# Go to parent directory
	os.chdir(os.path.dirname(os.path.realpath(__file__)) + '/..')

	# Create a solving instance of the model
	instance = dict()

	node_dict = {}


	# Load data to the instance
	load(instance, f'lcres/{pattern}.dot', 'pattern')
	load(instance, f'lcres/{target}.dot', 'target')

	load_position(instance, f'dat/{pos_pattern}.dot', 'pattern')
	load_position(instance, f'dat/{pos_target}.dot', 'target')

	if card: 
		card_p= cp_array(instance['lc_card_pattern'])
		card_t= cp_array(instance['lc_card_target'])

	preced_p_graph=instance['lc_preced_pattern']
	preced_t_graph=instance['lc_preced_target']


	pairs_p = [(i, elem) for i, sublist in enumerate(preced_p_graph) for elem in sublist]
	pairs_t = [(i, elem) for i, sublist in enumerate(preced_t_graph) for elem in sublist]


	I, V = model(**instance)


	timeout = 3600

	options = f"-t={timeout}s -s=all -trace -ev "

	#result = solve(sols=ALL, verbose=2)
	
	if(enable_timeout):
		solve (solver="ACE", options=options)
	else:
		result = solve(sols=ALL)
	ace_jar_path = "/home/etud/Bureau/projet/lib/python3.11/site-packages/pycsp3/solvers/ace/ACE-2.3.jar"


	if os.path.exists("bimodel.xml"):
		if(enable_timeout):
			command = ["java", "-jar", ace_jar_path, "bimodel.xml", "-s=all",   f"-t={timeout}s","-trace", "-ev"]
		else:
			command = ["java", "-jar", ace_jar_path, "bimodel.xml", "-s=all",   "-trace", "-ev"]
		try:
			result = subprocess.run(command, capture_output=True, text=True,timeout=timeout  )
			# Capture and print the solver's output
			# solver_output = result.stdout
			# print("Solver Output: uncomment to print")
			
			solver_output = result.stdout or ""
			if solver_output:
					#import_to_excel(solver_output)
					import_log_file(solver_output)


			if result.returncode != 0:
				print(f"ACE exited with error. ")
			else:
				print(f"ACE completed successfully. ")

		except subprocess.TimeoutExpired:
			print(f"ACE process timed out after {timeout} seconds. ")

	else:
		print("Error: bimodel.xml file not generated. Check your model definition.")


	
	
